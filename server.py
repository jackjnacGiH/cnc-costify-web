from flask import Flask, send_from_directory, request, jsonify, g
from flask_cors import CORS
import os
import sys
import time
import tempfile
import logging
import subprocess
import json
import urllib.request
import urllib.error
import socket
import io
import uuid
from logging.handlers import RotatingFileHandler

# Setup file logging for debugging packaged app (MOVED UP)
log_dir = os.path.join(os.path.expanduser('~'), 'Documents')
log_file = os.path.join(log_dir, 'CNC_Costify_AI_Log.txt')

# Configure root logger
root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)
# Remove existing handlers to avoid duplication
for handler in root_logger.handlers[:]:
    root_logger.removeHandler(handler)
# Add file handler
try:
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))
    root_logger.addHandler(file_handler)
except Exception as e:
    print(f"Failed to setup file logging: {e}")
# Add stream handler
stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))
root_logger.addHandler(stream_handler)

logging.info("Backend Server Started")
logging.info(f"Python Version: {sys.version}")
logging.info(f"Server Script Version: 3.0.0")

# Fix OCC Environment Variables (Critical for OneFile)
if getattr(sys, 'frozen', False):
    base_dir = sys._MEIPASS
    logging.info(f"Running in frozen mode. Base dir: {base_dir}")
    
    # Add base_dir to PATH to ensure DLLs are found
    os.environ['PATH'] = base_dir + os.pathsep + os.environ['PATH']
    logging.info(f"Added {base_dir} to PATH")

    # Check bundled resources
    occ_res_dir = os.path.join(base_dir, 'share', 'opencascade', 'resources')
    if os.path.exists(occ_res_dir):
        logging.info(f"Found bundled OCC resources at {occ_res_dir}")
        if 'CSF_UnitsDefinition' not in os.environ:
             units_file = os.path.join(occ_res_dir, 'UnitsAPI', 'Units.dat')
             if os.path.exists(units_file):
                 os.environ['CSF_UnitsDefinition'] = units_file
                 logging.info(f"Set CSF_UnitsDefinition to {units_file}")
             else:
                 logging.error(f"Units.dat not found at {units_file}")
        
        if 'CSF_STEPDefaults' not in os.environ:
             step_dir = os.path.join(occ_res_dir, 'XSTEPResource')
             if os.path.exists(step_dir):
                 os.environ['CSF_STEPDefaults'] = step_dir
                 logging.info(f"Set CSF_STEPDefaults to {step_dir}")

        # Set other necessary variables
        if 'CASROOT' not in os.environ:
             os.environ['CASROOT'] = os.path.dirname(occ_res_dir)

        if 'CSF_StandardDefaults' not in os.environ:
             os.environ['CSF_StandardDefaults'] = os.path.join(occ_res_dir, 'StdResource')
             os.environ['CSF_PluginDefaults'] = os.path.join(occ_res_dir, 'StdResource')
             os.environ['CSF_XCAFDefaults'] = os.path.join(occ_res_dir, 'StdResource')
        
        if 'CSF_ShadersDirectory' not in os.environ:
             os.environ['CSF_ShadersDirectory'] = os.path.join(occ_res_dir, 'Shaders')

    else:
        logging.warning(f"Bundled OCC resources NOT found at {occ_res_dir}")

try:
    from OCC.Core.STEPControl import STEPControl_Reader
    from OCC.Core.IFSelect import IFSelect_RetDone
    from OCC.Core.TopExp import TopExp_Explorer
    from OCC.Core.TopAbs import TopAbs_SOLID
    from OCC.Core.GProp import GProp_GProps
    from OCC.Core.BRepGProp import brepgprop_VolumeProperties, brepgprop_SurfaceProperties
    from OCC.Core.gp import gp_Ax1
    OCC_AVAILABLE = True
except Exception as e:
    logging.error(f"Failed to import OCC: {e}")
    OCC_AVAILABLE = False
    if getattr(sys, 'frozen', False):
        try:
             logging.info(f"Listing _MEIPASS content: {os.listdir(sys._MEIPASS)}")
        except Exception as list_err:
             logging.error(f"Failed to list _MEIPASS: {list_err}")

app = Flask(__name__)
CORS(app)

logging.info(f"OCC_AVAILABLE: {OCC_AVAILABLE}")
if OCC_AVAILABLE:
    logging.info(f"CSF_UnitsDefinition: {os.environ.get('CSF_UnitsDefinition', 'Not Set')}")
    if os.environ.get('CSF_UnitsDefinition'):
        logging.info(f"  File exists: {os.path.exists(os.environ['CSF_UnitsDefinition'])}")
    logging.info(f"CSF_UnitsLexicon: {os.environ.get('CSF_UnitsLexicon', 'Not Set')}")

# Resolve base directory for static files (supports PyInstaller onefile via _MEIPASS)
BASE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(__file__))

# Persistent calibration reference path
CALIBRATION_PATH = os.path.join(os.path.dirname(__file__), 'calibration_reference.json')

# --- AI Provider Config (persistent) ---
# Prefer storing config under ProgramData; fallback to LocalAppData; last resort next to script
try:
    _cfg_programdata = os.environ.get('PROGRAMDATA', r'C:\ProgramData')
except Exception:
    _cfg_programdata = r'C:\ProgramData'
try:
    CONFIG_DIR = os.path.join(_cfg_programdata, 'CNC Costify AI', 'config')
    os.makedirs(CONFIG_DIR, exist_ok=True)
except Exception:
    try:
        _localappdata = os.environ.get('LOCALAPPDATA') or os.path.join(os.path.expanduser('~'), 'AppData', 'Local')
    except Exception:
        _localappdata = os.path.join(os.path.expanduser('~'), 'AppData', 'Local')
    CONFIG_DIR = os.path.join(_localappdata, 'CNC Costify AI', 'config')
    try:
        os.makedirs(CONFIG_DIR, exist_ok=True)
    except Exception:
        CONFIG_DIR = os.path.join(os.path.dirname(__file__), 'config')
        os.makedirs(CONFIG_DIR, exist_ok=True)

AI_CONFIG_PATH = os.path.join(CONFIG_DIR, 'api_config.json')
AI_CONFIG_PATH_LEGACY = os.path.join(os.path.dirname(__file__), 'api_config.json')

# --- Runtime timeouts (tunable via environment) ---
try:
    GEMINI_TIMEOUT_SECS = int(os.environ.get('GEMINI_TIMEOUT_SECS') or '25')
except Exception:
    GEMINI_TIMEOUT_SECS = 25

def _mask_key(k: str) -> str:
    try:
        if not k:
            return ''
        k = str(k)
        if len(k) <= 8:
            return '*' * len(k)
        return f"{k[:4]}...{k[-4:]}"
    except Exception:
        return ''

def _load_ai_config() -> dict:
    cfg = {}
    loaded_from_legacy = False
    try:
        path_candidates = [AI_CONFIG_PATH, AI_CONFIG_PATH_LEGACY]
        for _p in path_candidates:
            if os.path.exists(_p):
                with open(_p, 'r', encoding='utf-8') as f:
                    cfg = json.load(f) or {}
                loaded_from_legacy = (_p == AI_CONFIG_PATH_LEGACY)
                break
    except Exception:
        try:
            app.logger.exception('Failed to read AI config')
        except Exception:
            pass
    # If read from legacy path, migrate to primary path when missing
    try:
        if loaded_from_legacy and not os.path.exists(AI_CONFIG_PATH):
            _save_ai_config(cfg)
    except Exception:
        try:
            app.logger.warning('Failed to migrate AI config from legacy path')
        except Exception:
            pass
    # defaults (no proxy Base URL, no Azure Speech)
    provider = cfg.get('provider') or os.environ.get('AI_PROVIDER') or 'gpt'
    openai = cfg.get('openai') or {}
    gemini = cfg.get('gemini') or {}
    cfg = {
        'provider': provider,
        'openai': {
            'api_key': openai.get('api_key') or os.environ.get('OPENAI_API_KEY') or '',
            'model': openai.get('model') or os.environ.get('OPENAI_MODEL') or 'gpt-3.5-turbo'
        },
        'gemini': {
            'api_key': gemini.get('api_key') or os.environ.get('GOOGLE_API_KEY') or os.environ.get('GEMINI_API_KEY') or '',
            'model': gemini.get('model') or os.environ.get('GEMINI_MODEL') or 'gemini-1.5-flash'
        }
    }
    return cfg

def _save_ai_config(cfg: dict) -> None:
    try:
        # Persist only provider/openai/gemini without base_url or azure_speech
        sanitized = {
            'provider': cfg.get('provider'),
            'openai': {
                'api_key': (cfg.get('openai') or {}).get('api_key') or '',
                'model': (cfg.get('openai') or {}).get('model') or 'gpt-3.5-turbo'
            },
            'gemini': {
                'api_key': (cfg.get('gemini') or {}).get('api_key') or '',
                'model': (cfg.get('gemini') or {}).get('model') or 'gemini-1.5-flash'
            }
        }
        try:
            with open(AI_CONFIG_PATH, 'w', encoding='utf-8') as f:
                json.dump(sanitized, f, ensure_ascii=False, indent=2)
        except Exception as e1:
            # Fallback to legacy path if primary write fails
            try:
                with open(AI_CONFIG_PATH_LEGACY, 'w', encoding='utf-8') as f:
                    json.dump(sanitized, f, ensure_ascii=False, indent=2)
                try:
                    app.logger.warning('AI config saved to legacy path due to write error on primary')
                except Exception:
                    pass
            except Exception as e2:
                try:
                    app.logger.exception(f'Failed to save AI config: {e1} | fallback: {e2}')
                except Exception:
                    pass
                raise
    except Exception:
        try:
            app.logger.exception('Failed to save AI config')
        except Exception:
            pass
        # Propagate to caller so UI can report failure
        raise

def _refresh_env_from_cfg(cfg: dict) -> None:
    try:
        os.environ['AI_PROVIDER'] = str(cfg.get('provider') or 'gpt')
        o = cfg.get('openai') or {}
        g_ = cfg.get('gemini') or {}
        os.environ['OPENAI_API_KEY'] = str(o.get('api_key') or '')
        os.environ['OPENAI_MODEL'] = str(o.get('model') or 'gpt-3.5-turbo')
        # Do not use proxy base URL
        os.environ.pop('OPENAI_BASE_URL', None)
        os.environ['GOOGLE_API_KEY'] = str(g_.get('api_key') or '')
        os.environ['GEMINI_API_KEY'] = os.environ['GOOGLE_API_KEY']
        os.environ['GEMINI_MODEL'] = str(g_.get('model') or 'gemini-1.5-flash')
        # Remove Azure Speech related envs
        os.environ.pop('AZURE_SPEECH_KEY', None)
        os.environ.pop('AZURE_SPEECH_REGION', None)
    except Exception:
        try:
            app.logger.exception('Failed to refresh env from cfg')
        except Exception:
            pass

# File logging suitable for No Console mode and Windows installer
# Prefer ProgramData to avoid write permission issues under Program Files
_programdata = os.environ.get('PROGRAMDATA', r'C:\ProgramData')
LOG_DIR = os.path.join(_programdata, 'CNC Costify AI', 'logs')
try:
    os.makedirs(LOG_DIR, exist_ok=True)
except Exception:
    # Fallback to local logs directory next to server.py
    LOG_DIR = os.path.join(os.path.dirname(__file__), 'logs')
    os.makedirs(LOG_DIR, exist_ok=True)

_log_path = os.path.join(LOG_DIR, 'server.log')
_handler = RotatingFileHandler(_log_path, maxBytes=2 * 1024 * 1024, backupCount=3, encoding='utf-8')
_formatter = logging.Formatter('%(asctime)s | %(levelname)s | pid=%(process)d | %(name)s | %(message)s')
_handler.setFormatter(_formatter)
_handler.setLevel(logging.INFO)
app.logger.addHandler(_handler)
app.logger.setLevel(logging.INFO)
_wz = logging.getLogger('werkzeug')
_wz.addHandler(_handler)
_wz.setLevel(logging.INFO)

# Request/Response logging with duration
@app.before_request
def _log_req():
    try:
        g._start = time.time()
    except Exception:
        g._start = None
    try:
        qs = request.query_string.decode('utf-8', 'ignore')
    except Exception:
        qs = ''
    app.logger.info(f"REQ {request.method} {request.path} qs='{qs}' len={request.content_length} ip={request.remote_addr}")
    # Attempt to log route matching info
    try:
        b = app.url_map.bind('')
        match = b.match(request.path, method=request.method)
        try:
            app.logger.info(f"ROUTE MATCH endpoint={match[0]} values={match[1]}")
        except Exception:
            app.logger.info(f"ROUTE MATCH raw={match}")
    except Exception as _ex:
        try:
            app.logger.info(f"ROUTE NO MATCH: {str(_ex)}")
        except Exception:
            pass

@app.after_request
def _log_res(resp):
    # Force /api/health 200 if any handler returned 404
    try:
        p = request.path.rstrip('/')
    except Exception:
        p = request.path
    if p == '/api/health' and request.method in ('GET', 'HEAD') and getattr(resp, 'status_code', 200) == 404:
        try:
            app.logger.info('AFTER_REQUEST override 404 /api/health -> 200')
        except Exception:
            pass
        from flask import Response
        import json as _json
        payload = {
            'ok': True,
            'status': 'running',
            'occ_available': bool(OCC_AVAILABLE),
            'time': time.time()
        }
        resp = Response(_json.dumps(payload), status=200, mimetype='application/json')
    try:
        started = getattr(g, '_start', None) or time.time()
        dur_ms = int((time.time() - started) * 1000)
    except Exception:
        dur_ms = -1
    app.logger.info(f"RES {resp.status_code} {request.method} {request.path} {dur_ms}ms")
    return resp

# Ensure /api/health responds 200 for GET/HEAD regardless of route resolution
@app.before_request
def _health_short_circuit():
    try:
        p = request.path.rstrip('/')
    except Exception:
        p = request.path
    if p == '/api/health' and request.method in ('GET', 'HEAD'):
        try:
            app.logger.info('SHORT CIRCUIT /api/health -> 200')
        except Exception:
            pass
        return jsonify({
            'ok': True,
            'status': 'running',
            'occ_available': bool(OCC_AVAILABLE),
            'time': time.time()
        })

@app.errorhandler(404)
def _health_404_override(e):
    try:
        p = request.path.rstrip('/')
    except Exception:
        p = request.path
    if p == '/api/health' and request.method in ('GET', 'HEAD'):
        try:
            app.logger.info('ERRORHANDLER 404 override /api/health -> 200')
        except Exception:
            pass
        return jsonify({
            'ok': True,
            'status': 'running',
            'occ_available': bool(OCC_AVAILABLE),
            'time': time.time()
        }), 200
    return e

@app.route('/')
def serve_html():
    # In packaged mode, do NOT serve the front-end at "/" to avoid browsers opening
    # Only the Electron app should display the UI. Return a minimal text in packaged mode.
    try:
        _is_frozen = getattr(sys, 'frozen', False)
    except Exception:
        _is_frozen = False
    if _is_frozen:
        from flask import Response
        return Response('CNC Costify AI backend running', mimetype='text/plain')
    return send_from_directory(BASE_DIR, 'CNC_Costify_AI_V6.html')

# CORS for front-end origins (5000 for npm, 5500 for local preview)
@app.after_request
def add_cors_headers(response):
    try:
        origin = request.headers.get('Origin', '')
    except Exception:
        origin = ''
    allowed = {
        'http://localhost:5000',
        'http://localhost:5500',
    }
    response.headers['Access-Control-Allow-Origin'] = origin if origin in allowed else '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET,POST,OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response


# Lightweight health-check endpoint used by dev auto-start logic
# Support both with/without trailing slash, and explicit HEAD
@app.route('/api/health', methods=['GET', 'HEAD', 'OPTIONS'])
@app.route('/api/health/', methods=['GET', 'HEAD', 'OPTIONS'])
def api_health():
    try:
        try:
            app.logger.info('ROUTE api_health invoked')
        except Exception:
            pass
        return jsonify({
            'ok': True,
            'status': 'running',
            'occ_available': bool(OCC_AVAILABLE),
            'time': time.time()
        })
    except Exception as e:
        app.logger.exception('api_health failed')
        return jsonify({'ok': False, 'error': str(e)}), 500

# --- AI Provider Config Endpoints (Flask) ---
@app.route('/api/config/ai', methods=['GET'])
def api_config_ai_get():
    try:
        cfg = _load_ai_config()
        out = {
            'ok': True,
            'provider': cfg.get('provider') or 'gpt',
            'openai': {
                'model': (cfg.get('openai') or {}).get('model') or 'gpt-3.5-turbo',
                'key_masked': _mask_key((cfg.get('openai') or {}).get('api_key') or '')
            },
            'gemini': {
                'model': (cfg.get('gemini') or {}).get('model') or 'gemini-1.5-flash',
                'key_masked': _mask_key((cfg.get('gemini') or {}).get('api_key') or '')
            }
        }
        return jsonify(out)
    except Exception as e:
        try:
            app.logger.exception('api_config_ai_get failed')
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/config/ai', methods=['POST'])
def api_config_ai_post():
    try:
        data = request.get_json(silent=True) or {}
        cfg = _load_ai_config()
        provider = data.get('provider')
        if provider:
            cfg['provider'] = str(provider).lower()
        o = cfg.get('openai') or {}
        g_ = cfg.get('gemini') or {}
        if isinstance(data.get('openaiApiKey'), str):
            o['api_key'] = data.get('openaiApiKey').strip()
        if isinstance(data.get('openaiModel'), str):
            o['model'] = data.get('openaiModel').strip()
        if isinstance(data.get('geminiApiKey'), str):
            g_['api_key'] = data.get('geminiApiKey').strip()
        if isinstance(data.get('geminiModel'), str):
            g_['model'] = data.get('geminiModel').strip()
        cfg['openai'] = o
        cfg['gemini'] = g_
        _save_ai_config(cfg)
        _refresh_env_from_cfg(cfg)
        return jsonify({
            'ok': True,
            'provider': cfg.get('provider'),
            'openai': {
                'model': o.get('model'),
                'key_masked': _mask_key(o.get('api_key') or '')
            },
            'gemini': {
                'model': g_.get('model'),
                'key_masked': _mask_key(g_.get('api_key') or '')
            }
        })
    except Exception as e:
        try:
            app.logger.exception('api_config_ai_post failed')
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

def _http_get(url: str, headers: dict = None, timeout: int = 6, retries: int = 0, backoff: float = 0.75):
    headers = headers or {}
    req = urllib.request.Request(url, headers=headers, method='GET')
    last_err = None
    attempt = 0
    while attempt <= max(0, int(retries)):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                code = resp.getcode()
                data = resp.read()
                return code, data.decode('utf-8', 'ignore')
        except urllib.error.HTTPError:
            # Propagate HTTP errors to caller to handle status-specific fallbacks
            raise
        except (urllib.error.URLError, socket.timeout) as e:
            last_err = e
            if attempt >= retries:
                raise
            try:
                time.sleep(backoff * (2 ** attempt))
            except Exception:
                pass
            attempt += 1
    # Should not reach here; raise last error if present
    if last_err:
        raise last_err
    raise RuntimeError('Unknown error during _http_get')

def _http_post(url: str, headers: dict = None, json_body: dict = None, timeout: int = 8, retries: int = 0, backoff: float = 0.75):
    headers = headers or {}
    data_bytes = None
    if json_body is not None:
        try:
            data_bytes = json.dumps(json_body).encode('utf-8')
        except Exception:
            data_bytes = None
        headers.setdefault('Content-Type', 'application/json')
    req = urllib.request.Request(url, headers=headers, method='POST')
    last_err = None
    attempt = 0
    while attempt <= max(0, int(retries)):
        try:
            with urllib.request.urlopen(req, data=data_bytes, timeout=timeout) as resp:
                code = resp.getcode()
                data = resp.read()
                return code, data.decode('utf-8', 'ignore')
        except urllib.error.HTTPError:
            # Propagate HTTP errors to caller to handle status-specific fallbacks
            raise
        except (urllib.error.URLError, socket.timeout) as e:
            last_err = e
            if attempt >= retries:
                raise
            try:
                time.sleep(backoff * (2 ** attempt))
            except Exception:
                pass
            attempt += 1
    if last_err:
        raise last_err
    raise RuntimeError('Unknown error during _http_post')

@app.route('/api/models/openai', methods=['GET'])
def api_models_openai():
    try:
        cfg = _load_ai_config()
        api_key = (cfg.get('openai') or {}).get('api_key') or ''
        base_url = 'https://api.openai.com/v1'
        if not api_key:
            return jsonify({'ok': False, 'error': 'OPENAI_API_KEY not configured'}), 400
        # Standard OpenAI REST (no proxy)
        url = f"{base_url}/models"
        try:
            code, text = _http_get(url, headers={'Authorization': f'Bearer {api_key}'} , timeout=8)
        except urllib.error.HTTPError as he:
            try:
                err_body = he.read().decode('utf-8', 'ignore')
            except Exception:
                err_body = str(he)
            try:
                obj = json.loads(err_body)
                msg = obj.get('error', {}).get('message') or obj.get('message') or err_body
            except Exception:
                msg = err_body or str(he)
            return jsonify({'ok': False, 'error': msg}), 500
        except urllib.error.URLError as ue:
            return jsonify({'ok': False, 'error': f'Network error: {ue.reason}'}), 500
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500
        try:
            obj = json.loads(text)
        except Exception:
            obj = {}
        data = obj.get('data') or []
        out = []
        for m in data:
            mid = str(m.get('id') or '')
            if not mid:
                continue
            can_chat = any(x in mid for x in ['gpt-4o', 'gpt-3.5-turbo', 'o3', 'gpt-4.1', 'gpt-4.1-mini'])
            out.append({'id': mid, 'displayName': mid, 'canChatCompletions': bool(can_chat)})
        # Fallback curated list if empty
        if not out:
            out = [
                {'id': 'gpt-4o-mini', 'displayName': 'GPT-4o mini', 'canChatCompletions': True},
                {'id': 'gpt-4o', 'displayName': 'GPT-4o', 'canChatCompletions': True},
                {'id': 'gpt-3.5-turbo', 'displayName': 'GPT-3.5 Turbo', 'canChatCompletions': True}
            ]
        return jsonify({'ok': True, 'models': out})
    except Exception as e:
        try:
            app.logger.exception('api_models_openai failed')
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

# --- STT (Speech-to-Text) via OpenAI Whisper ---
def _http_post_multipart(url: str, headers: dict, fields: dict, files: dict, timeout: int = 20):
    """Minimal multipart/form-data POST using urllib without external deps.
    fields: { name: value }
    files: { name: (filename, content_type, bytes) }
    Returns (code:int, text:str)
    """
    boundary = f"----CNCBoundary{uuid.uuid4().hex}"
    buf = io.BytesIO()
    # Write text fields
    for k, v in (fields or {}).items():
        buf.write((f"--{boundary}\r\n").encode('utf-8'))
        buf.write((f"Content-Disposition: form-data; name=\"{k}\"\r\n\r\n").encode('utf-8'))
        buf.write((str(v)).encode('utf-8'))
        buf.write(b"\r\n")
    # Write file parts
    for k, triple in (files or {}).items():
        try:
            filename, content_type, data = triple
        except Exception:
            # accept (filename, data) with auto content type
            if isinstance(triple, (list, tuple)) and len(triple) >= 2:
                filename, data = triple[0], triple[1]
                content_type = 'application/octet-stream'
            else:
                continue
        buf.write((f"--{boundary}\r\n").encode('utf-8'))
        buf.write((f"Content-Disposition: form-data; name=\"{k}\"; filename=\"{filename}\"\r\n").encode('utf-8'))
        buf.write((f"Content-Type: {content_type}\r\n\r\n").encode('utf-8'))
        if isinstance(data, str):
            data = data.encode('utf-8')
        buf.write(data)
        buf.write(b"\r\n")
    buf.write((f"--{boundary}--\r\n").encode('utf-8'))
    body = buf.getvalue()
    req_headers = dict(headers or {})
    req_headers['Content-Type'] = f"multipart/form-data; boundary={boundary}"
    req = urllib.request.Request(url, data=body, headers=req_headers, method='POST')
    with urllib.request.urlopen(req, timeout=timeout) as res:
        code = getattr(res, 'status', 200)
        text = res.read().decode('utf-8', 'ignore')
        return code, text

def _map_lang_for_whisper(lang: str) -> str:
    try:
        s = (lang or '').strip()
        if not s:
            return 'th'  # default
        s = s.lower()
        # map common locales
        if s in ('th', 'th-th'):
            return 'th'
        if s in ('en', 'en-us', 'en-gb'):
            return 'en'
        if '-' in s:
            return s.split('-')[0]
        return s
    except Exception:
        return 'th'

@app.route('/api/stt/transcribe', methods=['POST', 'OPTIONS'])
def api_stt_transcribe():
    # CORS for file-origin Electron UI
    if request.method == 'OPTIONS':
        from flask import Response
        resp = Response('', status=200)
        try:
            resp.headers['Access-Control-Allow-Origin'] = '*'
            resp.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        except Exception:
            pass
        return resp
    try:
        cfg = _load_ai_config()
        api_key = (cfg.get('openai') or {}).get('api_key') or os.environ.get('OPENAI_API_KEY') or ''
        model = 'whisper-1'
        if not api_key:
            return jsonify({'ok': False, 'error': 'OPENAI_API_KEY not configured'}), 400

        # Accept multipart upload
        file = None
        try:
            file = request.files.get('file')
        except Exception:
            file = None
        if not file:
            return jsonify({'ok': False, 'error': 'Missing audio file'}), 400
        data = file.read()
        fname = getattr(file, 'filename', None) or 'speech.webm'
        ctype = getattr(file, 'content_type', None) or 'audio/webm'
        lang = _map_lang_for_whisper(request.form.get('language') or '')

        # Call OpenAI transcription
        url = 'https://api.openai.com/v1/audio/transcriptions'
        headers = {'Authorization': f'Bearer {api_key}'}
        fields = {
            'model': model,
            'language': lang,
            'response_format': 'json'
        }
        files = {
            'file': (fname, ctype, data)
        }
        try:
            code, text = _http_post_multipart(url, headers=headers, fields=fields, files=files, timeout=30)
        except urllib.error.HTTPError as he:
            try:
                err_body = he.read().decode('utf-8', 'ignore')
            except Exception:
                err_body = str(he)
            try:
                obj = json.loads(err_body)
                msg = obj.get('error', {}).get('message') or obj.get('message') or err_body
            except Exception:
                msg = err_body or str(he)
            return jsonify({'ok': False, 'error': msg}), 500
        except urllib.error.URLError as ue:
            return jsonify({'ok': False, 'error': f'Network error: {ue.reason}'}), 500
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500

        try:
            obj = json.loads(text)
        except Exception:
            obj = {}
        out_text = obj.get('text') or obj.get('transcript') or ''
        resp = jsonify({'ok': True, 'text': out_text})
        try:
            resp.headers['Access-Control-Allow-Origin'] = '*'
        except Exception:
            pass
        return resp
    except Exception as e:
        try:
            app.logger.exception('api_stt_transcribe failed')
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/models/gemini', methods=['GET'])
def api_models_gemini():
    try:
        cfg = _load_ai_config()
        api_key = (cfg.get('gemini') or {}).get('api_key') or os.environ.get('GOOGLE_API_KEY') or os.environ.get('GEMINI_API_KEY') or ''
        if not api_key:
            return jsonify({'ok': False, 'error': 'GOOGLE_API_KEY/GEMINI_API_KEY not configured'}), 400
        api_base = 'https://generativelanguage.googleapis.com'
        def _list(ver: str):
            url = f"{api_base}/{ver}/models?key={api_key}"
            return _http_get(url, headers={}, timeout=GEMINI_TIMEOUT_SECS, retries=2)
        try:
            code, text = _list('v1beta')
        except urllib.error.HTTPError as he:
            # Fallback to v1 on NOT_FOUND/404
            try:
                body = he.read().decode('utf-8', 'ignore')
                obj = json.loads(body)
                status = (obj.get('error') or {}).get('status') or ''
                code_ = (obj.get('error') or {}).get('code') or he.code
            except Exception:
                status = ''
                code_ = getattr(he, 'code', 0)
            if status == 'NOT_FOUND' or code_ == 404:
                code, text = _list('v1')
            else:
                return jsonify({'ok': False, 'error': body or str(he)}), 500
        except urllib.error.URLError as ue:
            _r = getattr(ue, 'reason', ue)
            _msg = str(_r)
            if 'timed out' in _msg.lower():
                return jsonify({'ok': False, 'error': f'Timeout contacting Gemini ({GEMINI_TIMEOUT_SECS}s). Please try again.'}), 504
            return jsonify({'ok': False, 'error': f'Network error: {_msg}'}), 500
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500
        try:
            obj = json.loads(text)
        except Exception:
            obj = {}
        models = obj.get('models') or []
        out = []
        for m in models:
            mid = str(m.get('name') or m.get('id') or '')
            if not mid:
                continue
            disp = m.get('displayName') or mid
            methods = m.get('supportedGenerationMethods') or []
            can_gen = 'generateContent' in methods or 'create' in methods
            out.append({'id': mid.split('/')[-1], 'displayName': disp, 'canGenerateContent': bool(can_gen)})
        return jsonify({'ok': True, 'models': out})
    except Exception as e:
        try:
            app.logger.exception('api_models_gemini failed')
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500


# --- Chat endpoint (fixes 405 by providing POST /api/chat) ---
def _build_system_prompt(domain: str) -> str:
    d = str(domain or '').lower().strip()
    if d == 'density':
        return (
            'คุณคือผู้ช่วยด้านวัสดุ ช่วยตอบความหนาแน่น (Density) ของวัสดุที่ระบุ '
            'โดยอ้างอิงค่ามาตรฐานทั่วไป หน่วยหลักเป็น kg/m³ และให้คำแนะนำการใช้งานสั้นๆ '
            'หากไม่มั่นใจ ให้ระบุช่วงค่าที่พบได้และข้อควรระวังในการนำไปใช้'
        )
    elif 'วัสดุ' in d or d in ('material', 'spec', 'สเปควัตถุดิบ'):
        return (
            'คุณคือผู้ช่วยด้านสเปควัสดุ ให้ข้อมูลสเปค/คุณสมบัติ/การใช้งาน/วัสดุทดแทน '
            'ตอบกระชับ ใช้รายการหัวข้อสั้นๆ และเตือนข้อจำกัดเมื่อข้อมูลไม่แน่ชัด'
        )
    else:
        return (
            'คุณคือผู้ช่วยด้านวิศวกรรมและงาน CNC ตอบอย่างกระชับ ชัดเจน '
            'ให้ตัวอย่างและเหตุผลที่จำเป็นเท่านั้น'
        )

def _coerce_messages_for_openai(raw_messages):
    out = []
    for m in (raw_messages or []):
        try:
            role = str(m.get('role') or '').strip() or 'user'
            content = m.get('content')
            if isinstance(content, (list, dict)):
                content = json.dumps(content, ensure_ascii=False)
            content = str(content or '')
            out.append({'role': role, 'content': content})
        except Exception:
            continue
    return out

def _coerce_messages_for_gemini(raw_messages, system_text: str):
    # Gemini REST: contents: [{role: 'user'|'model', parts: [{text}]}]
    contents = []
    # Send system text as first user part to steer generation (portable across versions)
    if system_text:
        contents.append({'role': 'user', 'parts': [{'text': system_text}]})
    for m in (raw_messages or []):
        try:
            role = str(m.get('role') or '').strip().lower() or 'user'
            content = m.get('content')
            if isinstance(content, (list, dict)):
                content = json.dumps(content, ensure_ascii=False)
            content = str(content or '')
            r = 'user'
            if role == 'assistant':
                r = 'model'
            contents.append({'role': r, 'parts': [{'text': content}]})
        except Exception:
            continue
    return contents

# --- Lightweight material density DB and detector (parity with Express) ---
# Approximate density values in kg/m^3 and common synonyms
MATERIAL_DENSITY_DB = [
    { 'name': 'Steel (Carbon)', 'density_kg_m3': 7850, 'synonyms': ['steel', 'เหล็ก', 's45c', 's50c', 's55c', 'aisi 1045', 'aisi 1018'] },
    { 'name': 'เหล็ก SS400', 'density_kg_m3': 7850, 'synonyms': ['ss400', 'jis ss400', 'เหล็ก ss400'] },
    { 'name': 'Tool Steel SKD11', 'density_kg_m3': 7700, 'synonyms': ['skd11', 'tool steel', 'เอสเคดี11'] },
    { 'name': 'Stainless Steel 304', 'density_kg_m3': 8000, 'synonyms': ['stainless 304', 'stainless', 'สแตนเลส 304', 'sus304', 'ss304'] },
    { 'name': 'Stainless Steel 316', 'density_kg_m3': 8000, 'synonyms': ['stainless 316', 'สแตนเลส 316', 'sus316', 'ss316'] },
    { 'name': 'Aluminum 6061', 'density_kg_m3': 2700, 'synonyms': ['aluminum 6061', 'alu 6061', 'al6061', '6061', 'อลูมิเนียม 6061', 'อลูมิเนียม'] },
    { 'name': 'Aluminum 7075', 'density_kg_m3': 2810, 'synonyms': ['aluminum 7075', 'alu 7075', 'al7075', '7075', 'อลูมิเนียม 7075'] },
    { 'name': 'Brass', 'density_kg_m3': 8500, 'synonyms': ['brass', 'ทองเหลือง'] },
    { 'name': 'Copper', 'density_kg_m3': 8960, 'synonyms': ['copper', 'ทองแดง'] },
    { 'name': 'Titanium Ti-6Al-4V', 'density_kg_m3': 4430, 'synonyms': ['titanium', 'ti6al4v', 'ti-6al-4v', 'ไทเทเนียม'] },
    { 'name': 'PVC', 'density_kg_m3': 1400, 'synonyms': ['pvc', 'พีวีซี'] },
    { 'name': 'ABS', 'density_kg_m3': 1050, 'synonyms': ['abs', 'เอ บี เอส'] },
    { 'name': 'Nylon (PA6)', 'density_kg_m3': 1150, 'synonyms': ['nylon', 'pa6', 'ไนลอน'] },
    { 'name': 'POM (Delrin/Acetal)', 'density_kg_m3': 1410, 'synonyms': ['pom', 'delrin', 'acetal', 'พีโอเอ็ม', 'เดลริน'] },
    { 'name': 'PEEK', 'density_kg_m3': 1320, 'synonyms': ['peek', 'พีอีอีเค'] },
]

def detect_material_density(text):
    try:
        inp = str(text or '').lower()
    except Exception:
        inp = ''
    matches = []
    for entry in MATERIAL_DENSITY_DB:
        try:
            syns = entry.get('synonyms') or []
            for syn in syns:
                s = str(syn or '').lower()
                if not s:
                    continue
                if s in inp:
                    matches.append({
                        'material': entry.get('name'),
                        'density_kg_m3': entry.get('density_kg_m3'),
                        'synonym': syn,
                    })
                    break
        except Exception:
            continue
    return matches

@app.route('/api/material/density', methods=['POST', 'OPTIONS'])
def api_material_density():
    # Allow automatic OPTIONS by Flask; we still include for clarity with CORS headers
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    try:
        data = request.get_json(silent=True) or {}
        text = data.get('text')
        if not isinstance(text, str) or not text.strip():
            return jsonify({'ok': False, 'error': 'Missing text'}), 400
        matches = detect_material_density(text)
        return jsonify({'ok': True, 'matches': matches})
    except Exception as e:
        try:
            app.logger.exception('api_material_density failed')
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/chat', methods=['POST'])
def api_chat_post():
    try:
        data = request.get_json(silent=True) or {}
        provider = str(data.get('provider') or '').strip().lower() or None
        messages = data.get('messages') or []
        domain = data.get('domain') or 'general'

        # Load defaults from config if not provided
        cfg = _load_ai_config()
        if not provider:
            provider = str(cfg.get('provider') or 'gpt').lower()

        system_text = _build_system_prompt(domain)

        if provider in ('gpt', 'openai', 'chatgpt'):
            o = cfg.get('openai') or {}
            api_key = o.get('api_key') or os.environ.get('OPENAI_API_KEY') or ''
            model = o.get('model') or 'gpt-4o-mini'
            if not api_key:
                return jsonify({'ok': False, 'error': 'OPENAI_API_KEY not configured'}), 400
            base_url = 'https://api.openai.com/v1'
            url = f"{base_url}/chat/completions"
            msg = [{'role': 'system', 'content': system_text}] + _coerce_messages_for_openai(messages)
            payload = {
                'model': model,
                'messages': msg,
                'temperature': 0.2,
                'max_tokens': 800,
            }
            try:
                code, text = _http_post(url, headers={'Authorization': f'Bearer {api_key}'}, json_body=payload, timeout=12)
            except urllib.error.HTTPError as he:
                try:
                    body = he.read().decode('utf-8', 'ignore')
                except Exception:
                    body = str(he)
                try:
                    obj = json.loads(body)
                    msg = (obj.get('error') or {}).get('message') or obj.get('message') or body
                except Exception:
                    msg = body or str(he)
                return jsonify({'ok': False, 'error': msg}), 500
            except urllib.error.URLError as ue:
                return jsonify({'ok': False, 'error': f'Network error: {ue.reason}'}), 500
            except Exception as e:
                return jsonify({'ok': False, 'error': str(e)}), 500
            try:
                obj = json.loads(text)
            except Exception:
                obj = {}
            choices = obj.get('choices') or []
            content = ''
            if choices:
                c = choices[0] or {}
                msg_ = c.get('message') or {}
                content = msg_.get('content') or ''
            return jsonify({'ok': True, 'text': content or ''})

        elif provider in ('gemini', 'google'):
            g_ = cfg.get('gemini') or {}
            api_key = g_.get('api_key') or os.environ.get('GOOGLE_API_KEY') or os.environ.get('GEMINI_API_KEY') or ''
            model = g_.get('model') or 'gemini-1.5-flash'
            if not api_key:
                return jsonify({'ok': False, 'error': 'GOOGLE_API_KEY/GEMINI_API_KEY not configured'}), 400
            api_base = 'https://generativelanguage.googleapis.com'

            def _gen(ver: str):
                url = f"{api_base}/{ver}/models/{model}:generateContent?key={api_key}"
                contents = _coerce_messages_for_gemini(messages, system_text)
                payload = {'contents': contents}
                return _http_post(url, headers={}, json_body=payload, timeout=GEMINI_TIMEOUT_SECS, retries=2)

            try:
                code, text = _gen('v1beta')
            except urllib.error.HTTPError as he:
                # Fallback to v1 on NOT_FOUND/404
                try:
                    body = he.read().decode('utf-8', 'ignore')
                    obj = json.loads(body)
                    status = (obj.get('error') or {}).get('status') or ''
                    code_ = (obj.get('error') or {}).get('code') or he.code
                except Exception:
                    status = ''
                    code_ = getattr(he, 'code', 0)
                if status == 'NOT_FOUND' or code_ == 404:
                    code, text = _gen('v1')
                else:
                    return jsonify({'ok': False, 'error': body or str(he)}), 500
            except urllib.error.URLError as ue:
                _r = getattr(ue, 'reason', ue)
                _msg = str(_r)
                if 'timed out' in _msg.lower():
                    return jsonify({'ok': False, 'error': f'Timeout contacting Gemini ({GEMINI_TIMEOUT_SECS}s). Please try again.'}), 504
                return jsonify({'ok': False, 'error': f'Network error: {_msg}'}), 500
            except Exception as e:
                return jsonify({'ok': False, 'error': str(e)}), 500
            try:
                obj = json.loads(text)
            except Exception:
                obj = {}
            # Gemini returns candidates[0].content.parts[].text
            content = ''
            cands = obj.get('candidates') or []
            if cands:
                c = cands[0] or {}
                ct = c.get('content') or {}
                parts = ct.get('parts') or []
                if parts and isinstance(parts[0], dict):
                    content = parts[0].get('text') or ''
            return jsonify({'ok': True, 'text': content or ''})

        else:
            return jsonify({'ok': False, 'error': f'Unknown provider: {provider}'}), 400
    except Exception as e:
        try:
            app.logger.exception('api_chat_post failed')
        except Exception:
            pass
        return jsonify({'ok': False, 'error': str(e)}), 500


def compute_step_volume(filepath):
    reader = STEPControl_Reader()
    status = reader.ReadFile(filepath)
    if status != IFSelect_RetDone:
        raise RuntimeError('Failed to read STEP file')
    reader.TransferRoots()
    shape = reader.OneShape()
    total_volume = 0.0
    exp = TopExp_Explorer(shape, TopAbs_SOLID)
    has_solid = False
    while exp.More():
        solid = exp.Current()
        has_solid = True
        props = GProp_GProps()
        brepgprop_VolumeProperties(solid, props)
        total_volume += props.Mass()
        exp.Next()
    if not has_solid:
        props = GProp_GProps()
        brepgprop_VolumeProperties(shape, props)
        total_volume = props.Mass()
    return total_volume


def compute_step_area(filepath):
    reader = STEPControl_Reader()
    status = reader.ReadFile(filepath)
    if status != IFSelect_RetDone:
        raise RuntimeError('Failed to read STEP file')
    reader.TransferRoots()
    shape = reader.OneShape()
    total_area = 0.0
    exp = TopExp_Explorer(shape, TopAbs_SOLID)
    has_solid = False
    while exp.More():
        solid = exp.Current()
        has_solid = True
        props = GProp_GProps()
        brepgprop_SurfaceProperties(solid, props)
        total_area += props.Mass()
        exp.Next()
    if not has_solid:
        props = GProp_GProps()
        brepgprop_SurfaceProperties(shape, props)
        total_area = props.Mass()
    return total_area


def _load_and_clean_step_shape(filepath):
    """Load STEP file, filter for solids, and apply unit heuristic scaling if needed."""
    reader = STEPControl_Reader()
    status = reader.ReadFile(filepath)
    if status != IFSelect_RetDone:
        raise RuntimeError('Failed to read STEP file')
    reader.TransferRoots()
    shape = reader.OneShape()

    # Filter for solids
    try:
        from OCC.Core.TopoDS import TopoDS_Compound
        from OCC.Core.BRep import BRep_Builder
        from OCC.Core.TopExp import TopExp_Explorer
        from OCC.Core.TopAbs import TopAbs_SOLID
        
        solid_compound = TopoDS_Compound()
        builder = BRep_Builder()
        builder.MakeCompound(solid_compound)
        
        exp = TopExp_Explorer(shape, TopAbs_SOLID)
        has_solids = False
        while exp.More():
            has_solids = True
            builder.Add(solid_compound, exp.Current())
            exp.Next()
            
        if has_solids:
            shape = solid_compound
    except Exception:
        pass

    # Heuristic: Check for unit scaling issues (Meters vs Millimeters)
    try:
        from OCC.Core.Bnd import Bnd_Box
        from OCC.Core.BRepBndLib import brepbndlib_Add
        from OCC.Core.BRepBuilderAPI import BRepBuilderAPI_Transform
        from OCC.Core.gp import gp_Trsf, gp_Pnt
        
        bbox = Bnd_Box()
        brepbndlib_Add(shape, bbox)
        xmin, ymin, zmin, xmax, ymax, zmax = bbox.Get()
        dx = xmax - xmin
        dy = ymax - ymin
        dz = zmax - zmin
        max_dim = max(dx, dy, dz)
        
        logging.info(f"Unit heuristic check: Max dim = {max_dim:.2f}")

        # Threshold: 50,000 mm (50 meters). CNC parts are rarely this big.
        # If > 50m, and scaling by 0.001 brings it to > 1mm, assume it's a unit error.
        if max_dim > 50000.0:
            scaled_max = max_dim * 0.001
            if 1.0 < scaled_max < 50000.0:
                trsf = gp_Trsf()
                trsf.SetScale(gp_Pnt(0,0,0), 0.001)
                transformer = BRepBuilderAPI_Transform(shape, trsf)
                shape = transformer.Shape()
                logging.info(f"Unit heuristic: Scaled STEP shape by 0.001 (Max dim {max_dim:.2f} -> {scaled_max:.2f})")
            else:
                logging.info(f"Unit heuristic: Max dim {max_dim:.2f} too large but scaled {scaled_max:.2f} out of range")
        else:
            logging.info(f"Unit heuristic: Max dim {max_dim:.2f} within normal range")
    except Exception as e:
        logging.warning(f"Unit heuristic check failed: {e}")
        import traceback
        logging.warning(traceback.format_exc())

    return shape


def compute_step_stock_material(filepath, prefer_obb: bool = True):
    """Compute stock material dimensions and volume from a STEP file using OBB when possible.
    - Detect cylinder via cylindrical faces.
    - Use oriented bounding box (OBB) for dimensions and cylinder length mapping when available.
    - Fallback to axis-aligned bounding box (AABB) if OBB is not available.
    """
    import math
    # two-decimal rounding helper for stock size fields
    def _r2(x):
        try:
            return round(float(x), 2)
        except Exception:
            return x
    # Load and clean shape (includes solid filtering and unit heuristic)
    shape = _load_and_clean_step_shape(filepath)

    # Ensure triangulation to improve OBB quality on some OCCT builds
    try:
        from OCC.Core.BRepMesh import BRepMesh_IncrementalMesh
        # deflection ~0.5mm for typical parts; adjust as needed
        BRepMesh_IncrementalMesh(shape, 0.5, True, 0.5, True)
    except Exception:
        pass

    # Compute axis-aligned bounding box (AABB)
    from OCC.Core.Bnd import Bnd_Box
    try:
        # Prefer function form for broad compatibility
        from OCC.Core.BRepBndLib import brepbndlib_Add as _add_box
        aabb_box = Bnd_Box()
        _add_box(shape, aabb_box)
    except Exception:
        from OCC.Core.BRepBndLib import BRepBndLib
        aabb_box = Bnd_Box()
        BRepBndLib.Add(shape, aabb_box)
    xmin, ymin, zmin, xmax, ymax, zmax = aabb_box.Get()
    aabb_dx = float(xmax - xmin)
    aabb_dy = float(ymax - ymin)
    aabb_dz = float(zmax - zmin)

    # Attempt oriented bounding box (OBB)
    obb_available = False
    obb_dims = (aabb_dx, aabb_dy, aabb_dz)
    obb_axes = None
    obb_center = None
    try:
        from OCC.Core.Bnd import Bnd_OBB
        obb = Bnd_OBB()
        try:
            # Try class method first
            from OCC.Core.BRepBndLib import BRepBndLib
            # useTriangulation=True, useShapeTolerance=True for tighter OBB
            BRepBndLib.AddOBB(shape, obb, True, True)
        except Exception:
            # Try function form
            from OCC.Core.BRepBndLib import brepbndlib_AddOBB
            brepbndlib_AddOBB(shape, obb, True, True)
        # Extract half sizes and directions
        w = 2.0 * float(obb.XHSize())
        d = 2.0 * float(obb.YHSize())
        h = 2.0 * float(obb.ZHSize())
        # Directions (gp_Dir)
        xdir = obb.XDirection()
        ydir = obb.YDirection()
        zdir = obb.ZDirection()
        obb_axes = [
            (float(xdir.X()), float(xdir.Y()), float(xdir.Z())),
            (float(ydir.X()), float(ydir.Y()), float(ydir.Z())),
            (float(zdir.X()), float(zdir.Y()), float(zdir.Z())),
        ]
        obb_dims = (w, d, h)
        # Center (gp_Pnt)
        c = obb.Center()
        obb_center = (float(c.X()), float(c.Y()), float(c.Z()))
        obb_available = True
    except Exception:
        obb_available = False

    # Compute principal-axes-aligned extents using triangulated mesh points for robustness
    principal_extents = None
    principal_axes = None
    try:
        vprops = GProp_GProps()
        brepgprop_VolumeProperties(shape, vprops)
        pprops = vprops.PrincipalProperties()
        axis1 = pprops.FirstAxisOfInertia()
        axis2 = pprops.SecondAxisOfInertia()
        axis3 = pprops.ThirdAxisOfInertia()
        v1 = (float(axis1.X()), float(axis1.Y()), float(axis1.Z()))
        v2 = (float(axis2.X()), float(axis2.Y()), float(axis2.Z()))
        v3 = (float(axis3.X()), float(axis3.Y()), float(axis3.Z()))
        # Normalize principal axes to unit vectors
        def _norm(v):
            import math as _m
            n = _m.sqrt(v[0]*v[0] + v[1]*v[1] + v[2]*v[2]) or 1.0
            return (v[0]/n, v[1]/n, v[2]/n)
        v1 = _norm(v1)
        v2 = _norm(v2)
        v3 = _norm(v3)
        principal_axes = [v1, v2, v3]

        # Gather triangulation nodes from all faces
        from OCC.Core.TopAbs import TopAbs_FACE
        from OCC.Core.TopExp import TopExp_Explorer as _TopExpExplorer
        from OCC.Core.BRep import BRep_Tool
        from OCC.Core.TopLoc import TopLoc_Location
        from OCC.Core.gp import gp_Pnt

        mins = [float('inf'), float('inf'), float('inf')]
        maxs = [float('-inf'), float('-inf'), float('-inf')]
        any_node = False
        exp_faces2 = _TopExpExplorer(shape, TopAbs_FACE)
        while exp_faces2.More():
            face = exp_faces2.Current()
            loc = TopLoc_Location()
            tri = BRep_Tool.Triangulation(face, loc)
            if tri:
                nb = int(tri.NbNodes())
                # Location transform
                trsf = loc.Transformation()
                nodes = tri.Nodes()
                for i in range(1, nb + 1):
                    n = nodes.Value(i)
                    p = gp_Pnt(float(n.X()), float(n.Y()), float(n.Z()))
                    try:
                        p.Transform(trsf)
                    except Exception:
                        pass
                    x, y, z = float(p.X()), float(p.Y()), float(p.Z())
                    # Project to principal axes
                    u1 = x*v1[0] + y*v1[1] + z*v1[2]
                    u2 = x*v2[0] + y*v2[1] + z*v2[2]
                    u3 = x*v3[0] + y*v3[1] + z*v3[2]
                    if u1 < mins[0]: mins[0] = u1
                    if u2 < mins[1]: mins[1] = u2
                    if u3 < mins[2]: mins[2] = u3
                    if u1 > maxs[0]: maxs[0] = u1
                    if u2 > maxs[1]: maxs[1] = u2
                    if u3 > maxs[2]: maxs[2] = u3
                    any_node = True
            exp_faces2.Next()
        if any_node:
            w = float(maxs[0] - mins[0])
            d = float(maxs[1] - mins[1])
            h = float(maxs[2] - mins[2])
            principal_extents = (w, d, h)
    except Exception:
        principal_extents = None

    # Derive face-aligned orthonormal axes from dominant planar faces and compute OBB extents
    face_axes = None
    face_extents = None
    face_dominant_thickness = None
    try:
        from OCC.Core.TopAbs import TopAbs_FACE as _TA_FACE
        from OCC.Core.TopExp import TopExp_Explorer as _TE_Explorer
        from OCC.Core.BRepAdaptor import BRepAdaptor_Surface as _BRepAdaptorSurface
        from OCC.Core.GeomAbs import GeomAbs_Plane
        from OCC.Core.BRepGProp import brepgprop_SurfaceProperties as _surfProps

        def _normalize(v):
            import math as _m
            n = _m.sqrt(v[0]*v[0] + v[1]*v[1] + v[2]*v[2]) or 1.0
            return (v[0]/n, v[1]/n, v[2]/n)
        def _dot(a,b):
            return a[0]*b[0] + a[1]*b[1] + a[2]*b[2]
        def _cross(a,b):
            return (a[1]*b[2]-a[2]*b[1], a[2]*b[0]-a[0]*b[2], a[0]*b[1]-a[1]*b[0])

        # Cluster planar face normals by direction (treat +/- as same)
        groups = []  # list of dicts: {'n': rep_normal, 'area': total_area}
        cos_tol = 0.984  # ~10° tolerance for grouping
        exp_faces_f = _TE_Explorer(shape, _TA_FACE)
        while exp_faces_f.More():
            face = exp_faces_f.Current()
            try:
                surf = _BRepAdaptorSurface(face)
                if surf.GetType() == GeomAbs_Plane:
                    pln = surf.Plane()
                    ndir = pln.Axis().Direction()
                    n = _normalize((float(ndir.X()), float(ndir.Y()), float(ndir.Z())))
                    # canonicalize sign so n and -n group together
                    if n[2] < 0 or (n[2] == 0 and (n[1] < 0 or (n[1] == 0 and n[0] < 0))):
                        n = (-n[0], -n[1], -n[2])
                    # area of face
                    gp = GProp_GProps()
                    _surfProps(face, gp)
                    area = float(gp.Mass())
                    # assign to nearest group
                    assigned = False
                    for g in groups:
                        d = abs(_dot(n, g['n']))
                        if d >= cos_tol:
                            g['area'] += area
                            assigned = True
                            break
                    if not assigned:
                        groups.append({'n': n, 'area': area})
            except Exception:
                pass
            exp_faces_f.Next()

        # Pick z-axis as the most dominant planar normal
        groups_sorted = sorted(groups, key=lambda g: g['area'], reverse=True)
        if groups_sorted:
            z_axis = _normalize(groups_sorted[0]['n'])
            # pick x-axis as the most dominant group roughly perpendicular to z
            x_axis = None
            for g in groups_sorted[1:]:
                nd = _normalize(g['n'])
                if abs(_dot(nd, z_axis)) <= 0.20:  # ~78–102°
                    x_axis = nd
                    break
            # fallback: use a principal axis perpendicular to z
            if x_axis is None and principal_axes is not None:
                cands = sorted(principal_axes, key=lambda v: abs(_dot(v, z_axis)))
                if cands:
                    x_axis = _normalize(cands[0])
            # if still none, pick any orthogonal vector
            if x_axis is None:
                # choose a vector not colinear with z
                import math as _m
                tmp = (1.0, 0.0, 0.0) if abs(z_axis[0]) < 0.9 else (0.0, 1.0, 0.0)
                x_axis = _normalize(_cross(tmp, z_axis))
            # y = z × x to ensure orthonormal basis
            y_axis = _normalize(_cross(z_axis, x_axis))
            # re-orthogonalize x = y × z
            x_axis = _normalize(_cross(y_axis, z_axis))
            face_axes = [x_axis, y_axis, z_axis]

            # Project triangulation nodes onto face_axes to compute extents
            from OCC.Core.TopLoc import TopLoc_Location
            from OCC.Core.BRep import BRep_Tool
            from OCC.Core.gp import gp_Pnt
            mins = [float('inf'), float('inf'), float('inf')]
            maxs = [float('-inf'), float('-inf'), float('-inf')]
            any_node2 = False
            exp_faces_pts = _TE_Explorer(shape, _TA_FACE)
            while exp_faces_pts.More():
                face = exp_faces_pts.Current()
                loc = TopLoc_Location()
                tri = BRep_Tool.Triangulation(face, loc)
                if tri:
                    trsf = loc.Transformation()
                    nodes = tri.Nodes()
                    nb = int(tri.NbNodes())
                    for i in range(1, nb+1):
                        nnode = nodes.Value(i)
                        p = gp_Pnt(float(nnode.X()), float(nnode.Y()), float(nnode.Z()))
                        try:
                            p.Transform(trsf)
                        except Exception:
                            pass
                        x, y, z = float(p.X()), float(p.Y()), float(p.Z())
                        u1 = x*x_axis[0] + y*x_axis[1] + z*x_axis[2]
                        u2 = x*y_axis[0] + y*y_axis[1] + z*y_axis[2]
                        u3 = x*z_axis[0] + y*z_axis[1] + z*z_axis[2]
                        if u1 < mins[0]: mins[0] = u1
                        if u2 < mins[1]: mins[1] = u2
                        if u3 < mins[2]: mins[2] = u3
                        if u1 > maxs[0]: maxs[0] = u1
                        if u2 > maxs[1]: maxs[1] = u2
                        if u3 > maxs[2]: maxs[2] = u3
                        any_node2 = True
                exp_faces_pts.Next()
            if any_node2:
                face_extents = (float(maxs[0]-mins[0]), float(maxs[1]-mins[1]), float(maxs[2]-mins[2]))
                z_min_proj = float(mins[2])
                z_max_proj = float(maxs[2])

            # Measure dominant thickness using largest-area parallel planar faces aligned with z_axis
            try:
                from OCC.Core.BRepAdaptor import BRepAdaptor_Surface as _BRepAdaptorSurface2
                from OCC.Core.GeomAbs import GeomAbs_Plane as _GA_Plane2
                from OCC.Core.BRepGProp import brepgprop_SurfaceProperties as _surfProps2
                from OCC.Core.GProp import GProp_GProps as _GProps2
                import math as _m

                # Collect planar faces aligned with z_axis and compute plane offsets along z_axis
                faces_info = []  # (area, offset_s, sign)
                exp_faces_z = _TE_Explorer(shape, _TA_FACE)
                while exp_faces_z.More():
                    fz = exp_faces_z.Current()
                    try:
                        srf = _BRepAdaptorSurface2(fz)
                        if srf.GetType() == _GA_Plane2:
                            pln = srf.Plane()
                            ndir = pln.Axis().Direction()
                            nvec = _normalize((float(ndir.X()), float(ndir.Y()), float(ndir.Z())))
                            dotz = _dot(nvec, z_axis)
                            if abs(dotz) >= 0.985:  # ~10° within z_axis
                                # area
                                gp2 = _GProps2()
                                _surfProps2(fz, gp2)
                                area = float(gp2.Mass())
                                # plane reference point
                                p0 = pln.Location()
                                s_off = float(p0.X())*z_axis[0] + float(p0.Y())*z_axis[1] + float(p0.Z())*z_axis[2]
                                sign = 1 if dotz >= 0 else -1
                                faces_info.append((area, s_off, sign))
                    except Exception:
                        pass
                    exp_faces_z.Next()

                if faces_info:
                    # Group by offset with tolerance to merge coplanar patches
                    tol = 0.20  # mm
                    def _group_by_offset(entries):
                        groups = []  # (total_area, s_mean, members)
                        for (a, s, sign, face_obj) in entries:
                            placed = False
                            for g in groups:
                                if abs(s - g[1]) <= tol:
                                    # running mean weighted by area
                                    ta = g[0] + a
                                    g[1] = (g[1]*g[0] + s*a)/max(ta, 1e-9)
                                    g[0] = ta
                                    g[2].append((a, s, sign, face_obj))
                                    placed = True
                                    break
                            if not placed:
                                groups.append([a, s, [(a, s, sign, face_obj)]])
                        # sort by area desc
                        groups.sort(key=lambda x: x[0], reverse=True)
                        return groups

                    # split entries by normal sign (no face refs needed)
                    top_entries = [(a, s, sign, None) for (a, s, sign) in faces_info if sign > 0]
                    bot_entries = [(a, s, sign, None) for (a, s, sign) in faces_info if sign < 0]

                    # We need actual face objects; re-scan assigning faces to entries
                    if not top_entries or not bot_entries:
                        pass
                    else:
                        # Build groups
                        gtop = _group_by_offset(top_entries)
                        gbot = _group_by_offset(bot_entries)
                        if gtop and gbot:
                            # Prefer groups near the external extremes of z projection
                            ext_tol = 0.50  # mm window near extremes
                            # find top group closest to z_max_proj
                            top_area, top_s, top_members = min(
                                gtop,
                                key=lambda g: abs(g[1] - z_max_proj)
                            ) if 'z_max_proj' in locals() else gtop[0]
                            # find bottom group closest to z_min_proj
                            bot_area, bot_s, bot_members = min(
                                gbot,
                                key=lambda g: abs(g[1] - z_min_proj)
                            ) if 'z_min_proj' in locals() else gbot[0]

                            # If they are not close to extremes within tolerance, fallback to largest-area groups
                            if 'z_max_proj' in locals() and abs(top_s - z_max_proj) > ext_tol:
                                top_area, top_s, top_members = gtop[0]
                            if 'z_min_proj' in locals() and abs(bot_s - z_min_proj) > ext_tol:
                                bot_area, bot_s, bot_members = gbot[0]

                            if top_area > 0 and bot_area > 0:
                                face_dominant_thickness = float(top_s - bot_s)
                                if face_dominant_thickness <= 0:
                                    face_dominant_thickness = None
            except Exception:
                face_dominant_thickness = None
    except Exception:
        face_axes = None
        face_extents = None
        face_dominant_thickness = None

    # If principal extents are available, refine the thinnest dimension using large parallel planar faces
    refined_thickness = None
    thin_axis_index = None
    if principal_extents is not None and principal_axes is not None:
        try:
            # Identify thin axis
            dims_axes = list(principal_extents)
            thin_axis_index = int(min(range(3), key=lambda i: dims_axes[i]))
            thin_axis = principal_axes[thin_axis_index]

            # Scan faces to collect large planar faces aligned with thin axis
            from OCC.Core.TopAbs import TopAbs_FACE
            from OCC.Core.TopExp import TopExp_Explorer as _TopExpExplorer
            from OCC.Core.BRepAdaptor import BRepAdaptor_Surface as _BRepAdaptorSurface
            from OCC.Core.GeomAbs import GeomAbs_Plane
            from OCC.Core.BRepGProp import brepgprop_SurfaceProperties as _surfProps
            import math as _m

            aligned = []  # list of (area, s_avg, s_min, s_max, sign)
            aligned_faces = []  # list of dicts with face refs and metrics
            exp_faces3 = _TopExpExplorer(shape, TopAbs_FACE)
            while exp_faces3.More():
                face = exp_faces3.Current()
                try:
                    surf = _BRepAdaptorSurface(face)
                    if surf.GetType() == GeomAbs_Plane:
                        pln = surf.Plane()
                        # Plane normal
                        n = pln.Axis().Direction()
                        nvec = (float(n.X()), float(n.Y()), float(n.Z()))
                        dot = abs(nvec[0]*thin_axis[0] + nvec[1]*thin_axis[1] + nvec[2]*thin_axis[2])
                        # Accept faces whose normals align with thin axis within ~11.5° (cos>=0.98)
                        if dot >= 0.98:
                            # Face area
                            gp = GProp_GProps()
                            _surfProps(face, gp)
                            area = float(gp.Mass())
                            # Compute offsets from triangulation nodes projected along the face normal
                            from OCC.Core.TopLoc import TopLoc_Location as _TopLocLocation
                            from OCC.Core.BRep import BRep_Tool as _BRepTool
                            from OCC.Core.BRepMesh import BRepMesh_IncrementalMesh as _IncMesh
                            from OCC.Core.gp import gp_Pnt as _gp_Pnt

                            s_vals = []
                            try:
                                _loc = _TopLocLocation()
                                tri = _BRepTool.Triangulation(face, _loc)
                                if tri is None:
                                    try:
                                        _msh = _IncMesh(shape, 0.15)
                                        _msh.Perform()
                                        tri = _BRepTool.Triangulation(face, _loc)
                                    except Exception:
                                        tri = None
                                if tri is not None:
                                    trsf = _loc.Transformation()
                                    # Try Nodes() first
                                    try:
                                        nodes = tri.Nodes()
                                        nb = int(nodes.Length())
                                        for i in range(1, nb + 1):
                                            _p = nodes.Value(i)
                                            p = _gp_Pnt(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                            try:
                                                p.Transform(trsf)
                                            except Exception:
                                                pass
                                            sx, sy, sz = float(p.X()), float(p.Y()), float(p.Z())
                                            s_vals.append(sx*nvec[0] + sy*nvec[1] + sz*nvec[2])
                                    except Exception:
                                        try:
                                            nb = int(tri.NbNodes())
                                            for i in range(1, nb + 1):
                                                _p = tri.Node(i)
                                                p = _gp_Pnt(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                                try:
                                                    p.Transform(trsf)
                                                except Exception:
                                                    pass
                                                sx, sy, sz = float(p.X()), float(p.Y()), float(p.Z())
                                                s_vals.append(sx*nvec[0] + sy*nvec[1] + sz*nvec[2])
                                        except Exception:
                                            pass
                            except Exception:
                                pass
                            if s_vals:
                                s_min = float(min(s_vals))
                                s_max = float(max(s_vals))
                                s_avg = (s_min + s_max) * 0.5
                            else:
                                # Fallback to plane location if triangulation unavailable
                                loc = pln.Location()
                                s_avg = float(loc.X())*nvec[0] + float(loc.Y())*nvec[1] + float(loc.Z())*nvec[2]
                                s_min = s_avg
                                s_max = s_avg
                            true_dot = (nvec[0]*thin_axis[0] + nvec[1]*thin_axis[1] + nvec[2]*thin_axis[2])
                            sign = 1 if true_dot >= 0 else -1
                            aligned.append((area, s_avg, s_min, s_max, sign))
                            aligned_faces.append({
                                'face': face,
                                'area': area,
                                'sign': sign,
                                's_avg': s_avg,
                                's_min': s_min,
                                's_max': s_max,
                            })
                except Exception:
                    pass
                exp_faces3.Next()

            if len(aligned) >= 2:
                # Separate into top/bottom by normal sign relative to thin axis
                top = [(area, s_avg, s_min, s_max) for (area, s_avg, s_min, s_max, sign) in aligned if sign > 0]
                bottom = [(area, s_avg, s_min, s_max) for (area, s_avg, s_min, s_max, sign) in aligned if sign < 0]

                def _group_by_offset(items, tol=0.10):
                    groups = []  # (sum_area, avg_offset, min_offset, max_offset)
                    for area, s_avg, s_min, s_max in items:
                        placed = False
                        for i in range(len(groups)):
                            g_area, g_s, g_min, g_max = groups[i]
                            if abs(s_avg - g_s) <= tol:
                                new_area = g_area + area
                                new_s = (g_s * g_area + s_avg * area) / max(new_area, 1e-9)
                                groups[i] = (new_area, new_s, min(g_min, s_min), max(g_max, s_max))
                                placed = True
                                break
                        if not placed:
                            groups.append((area, s_avg, s_min, s_max))
                    groups.sort(key=lambda x: x[0], reverse=True)
                    return groups

                top_groups = _group_by_offset(top)
                bottom_groups = _group_by_offset(bottom)

                if top_groups and bottom_groups:
                    g_top_area, g_top_s, g_top_min, g_top_max = top_groups[0]
                    g_bot_area, g_bot_s, g_bot_min, g_bot_max = bottom_groups[0]
                    if g_top_area > 0 and g_bot_area > 0:
                        # Measure using extremes from dominant groups
                        delta = float(g_top_max - g_bot_min)
                        if delta >= 0.1:
                            refined_thickness = delta
                    else:
                        # Fallback: cross-pair among top 2 per side and pick max separation
                        candidates = []
                        for (a1, s1, m1, M1) in top_groups[:2]:
                            for (a2, s2, m2, M2) in bottom_groups[:2]:
                                if a1 > 0 and a2 > 0:
                                    candidates.append(float(M1 - m2))
                        if candidates:
                            refined_thickness = max(candidates)
                # Absolute fallback combining all planes if one side missing
                if refined_thickness is None:
                    # Use extremes across all aligned planar faces
                    all_groups = _group_by_offset([(a, s_avg, s_min, s_max) for (a, s_avg, s_min, s_max, _) in aligned])
                    mins = [g[2] for g in all_groups[:4] if len(all_groups) >= 1]
                    maxs = [g[3] for g in all_groups[:4] if len(all_groups) >= 1]
                    if mins and maxs:
                        refined_thickness = float(max(maxs) - min(mins))
                if refined_thickness <= 0:
                    refined_thickness = None

            # Exact/extracted distance between largest parallel outer faces (robust thickness)
            try:
                if aligned_faces:
                    max_area = max(e['area'] for e in aligned_faces)
                    # Keep only faces with substantial area (ignore small internal patches)
                    area_cut = max_area * 0.30
                    top_candidates = [e for e in aligned_faces if e['sign'] > 0 and e['area'] >= area_cut]
                    bot_candidates = [e for e in aligned_faces if e['sign'] < 0 and e['area'] >= area_cut]
                    if top_candidates and bot_candidates:
                        # Use extreme offsets along the thin axis across dominant faces
                        t_extreme = float(max(e['s_max'] for e in top_candidates) - min(e['s_min'] for e in bot_candidates))
                        if t_extreme > 0.0:
                            refined_thickness = t_extreme
                        else:
                            # Fallback to geometric distance in case extremes fail
                            try:
                                top_sel = sorted(top_candidates, key=lambda e: (e['area'], e['s_max']), reverse=True)[0]
                                bot_sel = sorted(bot_candidates, key=lambda e: (e['area'], -e['s_min']), reverse=True)[0]
                                from OCC.Core.BRepExtrema import BRepExtrema_DistShapeShape
                                dss = BRepExtrema_DistShapeShape(top_sel['face'], bot_sel['face'])
                                dss.Perform()
                                if dss.IsDone():
                                    dist_val = float(dss.Value())
                                    if dist_val > 0.0:
                                        refined_thickness = dist_val
                            except Exception:
                                pass
            except Exception:
                pass
            # Mesh-based reference points thickness (project triangulation nodes along thin axis)
            try:
                from OCC.Core.BRepMesh import BRepMesh_IncrementalMesh
                from OCC.Core.BRep import BRep_Tool as _BRepTool
                from OCC.Core.TopLoc import TopLoc_Location as _TopLocLocation
                from OCC.Core.TopAbs import TopAbs_FACE as _TA_FACE
                from OCC.Core.TopExp import TopExp_Explorer as _TE_Explorer
                from OCC.Core.BRepAdaptor import BRepAdaptor_Surface as _BRepAdaptorSurface
                from OCC.Core.GeomAbs import GeomAbs_Plane
                from OCC.Core.gp import gp_Pnt as _gp_Pnt

                # Triangulate with a moderate deflection to sample points
                try:
                    _mesher = BRepMesh_IncrementalMesh(shape, 0.20)
                    _mesher.Perform()
                except Exception:
                    pass

                exp_faces_m = _TE_Explorer(shape, _TA_FACE)
                min_s = None
                max_s = None
                while exp_faces_m.More():
                    f = exp_faces_m.Current()
                    try:
                        # Only consider planar faces whose normals align with thin axis
                        try:
                            _surf = _BRepAdaptorSurface(f)
                            if _surf.GetType() != GeomAbs_Plane:
                                raise RuntimeError('skip non-plane')
                            _pln = _surf.Plane()
                            _dn = _pln.Axis().Direction()
                            _nvec = (float(_dn.X()), float(_dn.Y()), float(_dn.Z()))
                            _dot = abs(_nvec[0]*thin_axis[0] + _nvec[1]*thin_axis[1] + _nvec[2]*thin_axis[2])
                            if _dot < 0.98:
                                raise RuntimeError('skip misaligned plane')
                        except Exception:
                            # If surface interrogation fails, skip this face from mesh projection
                            raise

                        loc = _TopLocLocation()
                        tri = _BRepTool.Triangulation(f, loc)
                        if tri is not None:
                            trsf = loc.Transformation()
                            # Prefer direct node access when available
                            try:
                                nodes = tri.Nodes()
                                nb = int(nodes.Length())
                                for i in range(1, nb + 1):
                                    _p = nodes.Value(i)
                                    p = _gp_Pnt(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                    try:
                                        p.Transform(trsf)
                                    except Exception:
                                        pass
                                    sx, sy, sz = float(p.X()), float(p.Y()), float(p.Z())
                                    s = sx*thin_axis[0] + sy*thin_axis[1] + sz*thin_axis[2]
                                    if min_s is None or s < min_s:
                                        min_s = s
                                    if max_s is None or s > max_s:
                                        max_s = s
                            except Exception:
                                # Fallback using NbNodes/Node accessors
                                try:
                                    nb = int(tri.NbNodes())
                                    for i in range(1, nb + 1):
                                        _p = tri.Node(i)
                                        p = _gp_Pnt(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                        try:
                                            p.Transform(trsf)
                                        except Exception:
                                            pass
                                        sx, sy, sz = float(p.X()), float(p.Y()), float(p.Z())
                                        s = sx*thin_axis[0] + sy*thin_axis[1] + sz*thin_axis[2]
                                        if min_s is None or s < min_s:
                                            min_s = s
                                        if max_s is None or s > max_s:
                                            max_s = s
                                except Exception:
                                    pass
                    except Exception:
                        pass
                    exp_faces_m.Next()

                if min_s is not None and max_s is not None and max_s > min_s:
                    mesh_thickness = float(max_s - min_s)
                    # Use mesh-based measurement only if planar grouping failed or disagrees notably
                    if refined_thickness is None or abs(mesh_thickness - refined_thickness) > 0.2:
                        refined_thickness = mesh_thickness
            except Exception:
                pass
        except Exception:
            refined_thickness = None

        # Histogram-based nominal thickness (robust to small local bosses/ribs)
        hist_thickness = None
        try:
            if thin_axis_index is not None:
                # Collect projected coordinates s along thin_axis from triangulation nodes of planar faces
                from OCC.Core.TopAbs import TopAbs_FACE as _TA_FACE_H
                from OCC.Core.TopExp import TopExp_Explorer as _TE_Explorer_H
                from OCC.Core.BRep import BRep_Tool as _BRepTool_H
                from OCC.Core.TopLoc import TopLoc_Location as _TopLocLocation_H
                from OCC.Core.BRepAdaptor import BRepAdaptor_Surface as _BRepAdaptorSurface_H
                from OCC.Core.GeomAbs import GeomAbs_Plane as _GeomAbs_Plane_H
                from OCC.Core.gp import gp_Pnt as _gp_Pnt_H
                import math as _m

                s_vals = []
                exp_faces_h = _TE_Explorer_H(shape, _TA_FACE_H)
                while exp_faces_h.More():
                    f = exp_faces_h.Current()
                    try:
                        _surf = _BRepAdaptorSurface_H(f)
                        if _surf.GetType() != _GeomAbs_Plane_H:
                            raise RuntimeError('skip non-plane')
                        _pln = _surf.Plane()
                        _dn = _pln.Axis().Direction()
                        _nvec = (float(_dn.X()), float(_dn.Y()), float(_dn.Z()))
                        _dot = abs(_nvec[0]*thin_axis[0] + _nvec[1]*thin_axis[1] + _nvec[2]*thin_axis[2])
                        if _dot < 0.98:
                            raise RuntimeError('skip misaligned plane')

                        loc = _TopLocLocation_H()
                        tri = _BRepTool_H.Triangulation(f, loc)
                        if tri is not None:
                            trsf = loc.Transformation()
                            try:
                                nodes = tri.Nodes()
                                nb = int(nodes.Length())
                                for i in range(1, nb + 1):
                                    _p = nodes.Value(i)
                                    p = _gp_Pnt_H(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                    try:
                                        p.Transform(trsf)
                                    except Exception:
                                        pass
                                    sx, sy, sz = float(p.X()), float(p.Y()), float(p.Z())
                                    s_vals.append(sx*thin_axis[0] + sy*thin_axis[1] + sz*thin_axis[2])
                            except Exception:
                                try:
                                    nb = int(tri.NbNodes())
                                    for i in range(1, nb + 1):
                                        _p = tri.Node(i)
                                        p = _gp_Pnt_H(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                        try:
                                            p.Transform(trsf)
                                        except Exception:
                                            pass
                                        sx, sy, sz = float(p.X()), float(p.Y()), float(p.Z())
                                        s_vals.append(sx*thin_axis[0] + sy*thin_axis[1] + sz*thin_axis[2])
                                except Exception:
                                    pass
                    except Exception:
                        pass
                    exp_faces_h.Next()

                if len(s_vals) >= 100:
                    s_min = float(min(s_vals))
                    s_max = float(max(s_vals))
                    span = s_max - s_min
                    if span > 1.0:
                        # Build histogram and smooth
                        bin_size = 0.25  # mm
                        bins = max(10, int(_m.ceil(span / bin_size)))
                        counts = [0] * bins
                        for s in s_vals:
                            idx = int(_m.floor((s - s_min) / bin_size))
                            if 0 <= idx < bins:
                                counts[idx] += 1
                        # 3-bin moving average smoothing
                        smooth = counts[:]
                        for i in range(bins):
                            acc = counts[i]
                            if i > 0:
                                acc += counts[i-1]
                            if i+1 < bins:
                                acc += counts[i+1]
                            smooth[i] = acc / 3.0
                        # Pick left and right peaks (near extremes) to avoid internal features
                        left_range = max(3, int(bins * 0.25))
                        right_start = max(left_range, int(bins * 0.75))
                        left_idx = max(range(left_range), key=lambda i: smooth[i])
                        right_idx = max(range(right_start, bins), key=lambda i: smooth[i])
                        s_left = s_min + (left_idx + 0.5) * bin_size
                        s_right = s_min + (right_idx + 0.5) * bin_size
                        cand = abs(s_right - s_left)
                        if cand > 0.5:
                            hist_thickness = cand
        except Exception:
            hist_thickness = None

    # Detect cylinder by scanning faces
    from OCC.Core.TopAbs import TopAbs_FACE
    from OCC.Core.TopExp import TopExp_Explorer
    from OCC.Core.BRepAdaptor import BRepAdaptor_Surface
    from OCC.Core.GeomAbs import GeomAbs_Cylinder

    exp_faces = TopExp_Explorer(shape, TopAbs_FACE)
    cylinder_radius = None
    cylinder_axis_dir = None
    cylinder_area = 0.0
    # Compute total surface area
    try:
        gp_total = GProp_GProps()
        brepgprop_SurfaceProperties(shape, gp_total)
        total_surface_area = float(gp_total.Mass())
    except Exception:
        total_surface_area = None

    # Group cylindrical faces by axis, then choose cluster with maximum radius
    cyl_faces = []  # [{'radius': r, 'axis': (ux,uy,uz), 'area': a, 'axis_pt': (x,y,z)}]
    while exp_faces.More():
        shape_face = exp_faces.Current()
        try:
            surf = BRepAdaptor_Surface(shape_face)
            if surf.GetType() == GeomAbs_Cylinder:
                cyl = surf.Cylinder()
                r = float(cyl.Radius())
                ax1 = cyl.Axis()
                d = ax1.Direction()
                # normalize axis direction
                ux, uy, uz = float(d.X()), float(d.Y()), float(d.Z())
                norm = (ux*ux + uy*uy + uz*uz) ** 0.5 or 1.0
                axu = (ux / norm, uy / norm, uz / norm)
                loc = ax1.Location()
                axis_pt = (float(loc.X()), float(loc.Y()), float(loc.Z()))
                # face area
                gp_face = GProp_GProps()
                brepgprop_SurfaceProperties(shape_face, gp_face)
                f_area = float(gp_face.Mass())
                cyl_faces.append({'radius': r, 'axis': axu, 'area': f_area, 'axis_pt': axis_pt})
        except Exception:
            pass
        exp_faces.Next()

    # Cluster faces by axis similarity (ignore sign), aggregate area, pick max radius per cluster
    clusters = []  # [{'rep_axis': (x,y,z), 'total_area': a, 'max_radius': r}]
    def _dot(a, b):
        return a[0]*b[0] + a[1]*b[1] + a[2]*b[2]
    for f in cyl_faces:
        placed = False
        for c in clusters:
            # axis similarity: allow opposite directions as same axis
            sim = abs(_dot(f['axis'], c['rep_axis']))
            if sim >= 0.98:
                c['total_area'] += f['area']
                if f['radius'] > c['max_radius']:
                    c['max_radius'] = f['radius']
                placed = True
                break
        if not placed:
            clusters.append({'rep_axis': f['axis'], 'total_area': f['area'], 'max_radius': f['radius']})

    cylinder_axis_point = None
    if clusters:
        # choose cluster with largest radius; outer stock should dominate
        best = sorted(clusters, key=lambda c: c['max_radius'], reverse=True)[0]
        cylinder_axis_dir = best['rep_axis']
        cylinder_area = best['total_area']
        # pick axis point from face with largest radius in this cluster
        import math as _m
        tol = 1e-6
        candidates = [f for f in cyl_faces if abs(_dot(f['axis'], cylinder_axis_dir)) >= 0.98]
        if candidates:
            best_face = sorted(candidates, key=lambda f: f['radius'], reverse=True)[0]
            cylinder_radius = best_face['radius']
            cylinder_axis_point = best_face.get('axis_pt')
        else:
            cylinder_radius = best['max_radius']

    result = {
        'bbox_mm': {
            'width_mm': aabb_dx,
            'depth_mm': aabb_dy,
            'height_mm': aabb_dz,
        }
    }
    if obb_available:
        result['obb_mm'] = {
            'width_mm': obb_dims[0],
            'depth_mm': obb_dims[1],
            'height_mm': obb_dims[2],
        }

    if cylinder_radius is not None and cylinder_axis_dir is not None:
        # Measure cylinder length by projecting triangulation nodes along cylinder axis
        import math as _m
        length_mm = None
        try:
            from OCC.Core.TopAbs import TopAbs_FACE as _TAF
            from OCC.Core.TopExp import TopExp_Explorer as _TE
            from OCC.Core.BRep import BRep_Tool as _BT
            from OCC.Core.TopLoc import TopLoc_Location as _TL
            from OCC.Core.gp import gp_Pnt as _P
            # normalize axis
            axn = cylinder_axis_dir
            n = _m.sqrt(axn[0]*axn[0] + axn[1]*axn[1] + axn[2]*axn[2]) or 1.0
            axu = (axn[0]/n, axn[1]/n, axn[2]/n)
            min_s = float('inf')
            max_s = float('-inf')
            exp_len = _TE(shape, _TAF)
            while exp_len.More():
                f = exp_len.Current()
                loc = _TL()
                tri = _BT.Triangulation(f, loc)
                if tri:
                    trsf = loc.Transformation()
                    # Prefer Nodes() when available
                    try:
                        nodes = tri.Nodes()
                        nb = int(nodes.Length())
                        for i in range(1, nb+1):
                            _p = nodes.Value(i)
                            p = _P(float(_p.X()), float(_p.Y()), float(_p.Z()))
                            try:
                                p.Transform(trsf)
                            except Exception:
                                pass
                            x, y, z = float(p.X()), float(p.Y()), float(p.Z())
                            s = x*axu[0] + y*axu[1] + z*axu[2]
                            if s < min_s: min_s = s
                            if s > max_s: max_s = s
                    except Exception:
                        try:
                            nb = int(tri.NbNodes())
                            for i in range(1, nb+1):
                                _p = tri.Node(i)
                                p = _P(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                try:
                                    p.Transform(trsf)
                                except Exception:
                                    pass
                                x, y, z = float(p.X()), float(p.Y()), float(p.Z())
                                s = x*axu[0] + y*axu[1] + z*axu[2]
                                if s < min_s: min_s = s
                                if s > max_s: max_s = s
                        except Exception:
                            pass
                exp_len.Next()
            if max_s > min_s and max_s < float('inf') and min_s > float('-inf'):
                length_mm = float(max_s - min_s)
        except Exception:
            length_mm = None

        # Fallback length from OBB/AABB mapping if projection failed
        if length_mm is None:
            if prefer_obb and obb_available and obb_axes:
                dots = [
                    abs(cylinder_axis_dir[0]*obb_axes[0][0] + cylinder_axis_dir[1]*obb_axes[0][1] + cylinder_axis_dir[2]*obb_axes[0][2]),
                    abs(cylinder_axis_dir[0]*obb_axes[1][0] + cylinder_axis_dir[1]*obb_axes[1][1] + cylinder_axis_dir[2]*obb_axes[1][2]),
                    abs(cylinder_axis_dir[0]*obb_axes[2][0] + cylinder_axis_dir[1]*obb_axes[2][1] + cylinder_axis_dir[2]*obb_axes[2][2]),
                ]
                axis_index = int(dots.index(max(dots)))
                length_mm = obb_dims[axis_index]
            else:
                ax = [abs(cylinder_axis_dir[0]), abs(cylinder_axis_dir[1]), abs(cylinder_axis_dir[2])]
                axis_index = int(ax.index(max(ax)))
                length_mm = [aabb_dx, aabb_dy, aabb_dz][axis_index]

        # Consider circular edges perpendicular to axis to capture largest steps/rims
        try:
            from OCC.Core.TopAbs import TopAbs_EDGE as _TAE
            from OCC.Core.TopExp import TopExp_Explorer as _TEE
            from OCC.Core.BRepAdaptor import BRepAdaptor_Curve as _BAC
            from OCC.Core.GeomAbs import GeomAbs_Circle as _GA_Circle
            # Centre of mass as axis point reference
            try:
                _vprops = GProp_GProps()
                brepgprop_VolumeProperties(shape, _vprops)
                _com = _vprops.CentreOfMass()
                com_pt = (float(_com.X()), float(_com.Y()), float(_com.Z()))
            except Exception:
                com_pt = (0.0, 0.0, 0.0)
            # unit axis
            import math as _m
            _n = _m.sqrt(cylinder_axis_dir[0]**2 + cylinder_axis_dir[1]**2 + cylinder_axis_dir[2]**2) or 1.0
            _axu = (cylinder_axis_dir[0]/_n, cylinder_axis_dir[1]/_n, cylinder_axis_dir[2]/_n)
            max_circle_r = None
            exp_edges = _TEE(shape, _TAE)
            while exp_edges.More():
                e = exp_edges.Current()
                try:
                    c = _BAC(e)
                    if c.GetType() == _GA_Circle:
                        circ = c.Circle()
                        ax = circ.Axis().Direction()
                        nvec = (float(ax.X()), float(ax.Y()), float(ax.Z()))
                        # Circle plane normal should align with cylinder axis (allow tolerance)
                        dot = abs(nvec[0]*cylinder_axis_dir[0] + nvec[1]*cylinder_axis_dir[1] + nvec[2]*cylinder_axis_dir[2])
                        if dot >= 0.90:
                            # Check circle centre proximity to axis line
                            cen = circ.Location()
                            pc = (float(cen.X()), float(cen.Y()), float(cen.Z()))
                            # distance from point to axis through COM: |(pc-com) x axu|
                            vx = pc[0] - com_pt[0]
                            vy = pc[1] - com_pt[1]
                            vz = pc[2] - com_pt[2]
                            cx = vy*_axu[2] - vz*_axu[1]
                            cy = vz*_axu[0] - vx*_axu[2]
                            cz = vx*_axu[1] - vy*_axu[0]
                            dist = _m.sqrt(cx*cx + cy*cy + cz*cz)
                            if dist <= 10.0:  # within 10mm of axis
                                r = float(circ.Radius())
                                if (max_circle_r is None) or (r > max_circle_r):
                                    max_circle_r = r
                except Exception:
                    pass
                exp_edges.Next()
        except Exception:
            max_circle_r = None

        # Mesh-based largest coaxial radius to axis line (true-geometry measure)
        mesh_max_r = None
        try:
            from OCC.Core.BRepMesh import BRepMesh_IncrementalMesh as _IncMesh_R
            from OCC.Core.BRep import BRep_Tool as _BRepTool_R
            from OCC.Core.TopLoc import TopLoc_Location as _TopLoc_R
            from OCC.Core.TopAbs import TopAbs_FACE as _TAF_R
            from OCC.Core.TopExp import TopExp_Explorer as _TE_R
            from OCC.Core.gp import gp_Pnt as _P_R
            # axis point fallback
            P0 = cylinder_axis_point if cylinder_axis_point is not None else com_pt
            ux, uy, uz = _axu
            # triangulate moderately
            try:
                _mesher_r = _IncMesh_R(shape, 0.20)
                _mesher_r.Perform()
            except Exception:
                pass
            exp_r = _TE_R(shape, _TAF_R)
            while exp_r.More():
                f = exp_r.Current()
                loc = _TopLoc_R()
                tri = _BRepTool_R.Triangulation(f, loc)
                if tri:
                    trsf = loc.Transformation()
                    # prefer Nodes when available
                    try:
                        nodes = tri.Nodes()
                        nb = int(nodes.Length())
                        for i in range(1, nb+1):
                            _p = nodes.Value(i)
                            p = _P_R(float(_p.X()), float(_p.Y()), float(_p.Z()))
                            try:
                                p.Transform(trsf)
                            except Exception:
                                pass
                            x, y, z = float(p.X()), float(p.Y()), float(p.Z())
                            vx, vy, vz = x - P0[0], y - P0[1], z - P0[2]
                            # radial vector component: v - (v·u)u
                            dotv = vx*ux + vy*uy + vz*uz
                            rx = vx - dotv*ux
                            ry = vy - dotv*uy
                            rz = vz - dotv*uz
                            rlen = _m.sqrt(rx*rx + ry*ry + rz*rz)
                            if (mesh_max_r is None) or (rlen > mesh_max_r):
                                mesh_max_r = rlen
                    except Exception:
                        try:
                            nb = int(tri.NbNodes())
                            for i in range(1, nb+1):
                                _p = tri.Node(i)
                                p = _P_R(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                try:
                                    p.Transform(trsf)
                                except Exception:
                                    pass
                                x, y, z = float(p.X()), float(p.Y()), float(p.Z())
                                vx, vy, vz = x - P0[0], y - P0[1], z - P0[2]
                                dotv = vx*ux + vy*uy + vz*uz
                                rx = vx - dotv*ux
                                ry = vy - dotv*uy
                                rz = vz - dotv*uz
                                rlen = _m.sqrt(rx*rx + ry*ry + rz*rz)
                                if (mesh_max_r is None) or (rlen > mesh_max_r):
                                    mesh_max_r = rlen
                        except Exception:
                            pass
                exp_r.Next()
        except Exception:
            mesh_max_r = None

        # Radial profile along axis: slice nodes by s (projection along axis), pick per-slice max radius
        profile_max_r = None
        profile_bins_debug = None
        profile_meta = None
        try:
            from OCC.Core.BRepMesh import BRepMesh_IncrementalMesh as _IncMesh_P
            from OCC.Core.BRep import BRep_Tool as _BRepTool_P
            from OCC.Core.TopLoc import TopLoc_Location as _TopLoc_P
            from OCC.Core.TopAbs import TopAbs_FACE as _TAF_P
            from OCC.Core.TopExp import TopExp_Explorer as _TE_P
            from OCC.Core.gp import gp_Pnt as _P_P
            # axis ref
            P0p = cylinder_axis_point if cylinder_axis_point is not None else com_pt
            ux, uy, uz = _axu
            # triangulate
            try:
                _mesher_p = _IncMesh_P(shape, 0.20)
                _mesher_p.Perform()
            except Exception:
                pass
            # First pass to get s-range
            s_min = None
            s_max = None
            exp_p = _TE_P(shape, _TAF_P)
            nodes_cache = []
            while exp_p.More():
                f = exp_p.Current()
                loc = _TopLoc_P()
                tri = _BRepTool_P.Triangulation(f, loc)
                if tri:
                    trsf = loc.Transformation()
                    try:
                        nodes = tri.Nodes()
                        nb = int(nodes.Length())
                        for i in range(1, nb+1):
                            _p = nodes.Value(i)
                            p = _P_P(float(_p.X()), float(_p.Y()), float(_p.Z()))
                            try:
                                p.Transform(trsf)
                            except Exception:
                                pass
                            x, y, z = float(p.X()), float(p.Y()), float(p.Z())
                            vx, vy, vz = x - P0p[0], y - P0p[1], z - P0p[2]
                            s = vx*ux + vy*uy + vz*uz
                            nodes_cache.append((x, y, z, s))
                            if s_min is None or s < s_min: s_min = s
                            if s_max is None or s > s_max: s_max = s
                    except Exception:
                        try:
                            nb = int(tri.NbNodes())
                            for i in range(1, nb+1):
                                _p = tri.Node(i)
                                p = _P_P(float(_p.X()), float(_p.Y()), float(_p.Z()))
                                try:
                                    p.Transform(trsf)
                                except Exception:
                                    pass
                                x, y, z = float(p.X()), float(p.Y()), float(p.Z())
                                vx, vy, vz = x - P0p[0], y - P0p[1], z - P0p[2]
                                s = vx*ux + vy*uy + vz*uz
                                nodes_cache.append((x, y, z, s))
                                if s_min is None or s < s_min: s_min = s
                                if s_max is None or s > s_max: s_max = s
                        except Exception:
                            pass
                exp_p.Next()
            # Build bins and compute per-bin max radius
            if s_min is not None and s_max is not None and s_max > s_min:
                import math as _m
                span = s_max - s_min
                # default finer resolution; allow override via query args
                bin_size = 0.25  # mm
                try:
                    from flask import request as _req
                    _bin_mm_arg = _req.args.get('profile_bin_mm', None)
                    _slices_arg = _req.args.get('profile_slices', None)
                except Exception:
                    _bin_mm_arg = None
                    _slices_arg = None
                if _bin_mm_arg:
                    try:
                        bval = float(_bin_mm_arg)
                        if bval > 0.05:
                            bin_size = bval
                    except Exception:
                        pass
                bins = max(5, int(_m.ceil(span / bin_size)))
                if _slices_arg:
                    try:
                        sval = int(_slices_arg)
                        if sval >= 8:
                            bins = sval
                            bin_size = span / float(bins)
                    except Exception:
                        pass
                max_r_bins = [0.0] * bins
                for (x, y, z, s) in nodes_cache:
                    vx, vy, vz = x - P0p[0], y - P0p[1], z - P0p[2]
                    dotv = vx*ux + vy*uy + vz*uz
                    rx = vx - dotv*ux
                    ry = vy - dotv*uy
                    rz = vz - dotv*uz
                    rlen = _m.sqrt(rx*rx + ry*ry + rz*rz)
                    idx = int(_m.floor((s - s_min) / bin_size))
                    if 0 <= idx < bins:
                        if rlen > max_r_bins[idx]:
                            max_r_bins[idx] = rlen
                # pick global max across slices
                profile_max_r = max(max_r_bins) if max_r_bins else None
                # build debug profile
                try:
                    limit = min(bins, 256)
                    profile_bins_debug = []
                    for i in range(limit):
                        s_center = s_min + (i + 0.5) * bin_size
                        r = float(max_r_bins[i])
                        profile_bins_debug.append({'s_mm': s_center, 'r_mm': r, 'd_mm': 2.0 * r})
                    profile_meta = {'bin_mm': float(bin_size), 's_min_mm': float(s_min), 's_max_mm': float(s_max)}
                except Exception:
                    pass
        except Exception:
            profile_max_r = None

        diameter_mm = 2.0 * cylinder_radius
        # prefer true-geometry circle edges if larger
        if max_circle_r is not None and (2.0 * max_circle_r) > diameter_mm:
            diameter_mm = 2.0 * max_circle_r
        # else prefer mesh-based largest coaxial radius if larger
        if mesh_max_r is not None and (2.0 * mesh_max_r) > diameter_mm:
            diameter_mm = 2.0 * mesh_max_r
        # else prefer radial-profile max if larger
        if profile_max_r is not None and (2.0 * profile_max_r) > diameter_mm:
            diameter_mm = 2.0 * profile_max_r
        # Validate cylinder: either dominant cylinder area or dims close to diameter
        area_frac_ok = False
        try:
            if total_surface_area and total_surface_area > 0:
                area_frac = cylinder_area / total_surface_area
                area_frac_ok = (area_frac >= 0.30)
        except Exception:
            area_frac_ok = False

        perp_dims = None
        if prefer_obb and obb_available and obb_axes:
            # Determine perpendicular OBB dims
            dots = [
                abs(cylinder_axis_dir[0]*obb_axes[0][0] + cylinder_axis_dir[1]*obb_axes[0][1] + cylinder_axis_dir[2]*obb_axes[0][2]),
                abs(cylinder_axis_dir[0]*obb_axes[1][0] + cylinder_axis_dir[1]*obb_axes[1][1] + cylinder_axis_dir[2]*obb_axes[1][2]),
                abs(cylinder_axis_dir[0]*obb_axes[2][0] + cylinder_axis_dir[1]*obb_axes[2][1] + cylinder_axis_dir[2]*obb_axes[2][2]),
            ]
            axis_index = int(dots.index(max(dots)))
            perp_dims = [obb_dims[i] for i in range(3) if i != axis_index]
        else:
            ax = [abs(cylinder_axis_dir[0]), abs(cylinder_axis_dir[1]), abs(cylinder_axis_dir[2])]
            axis_index = int(ax.index(max(ax)))
            perp_dims = [[aabb_dx, aabb_dy, aabb_dz][i] for i in range(3) if i != axis_index]

        rel_tol = 0.10
        abs_tol = 2.0
        valid_perp = all(abs(p - diameter_mm) <= max(abs_tol, rel_tol * max(diameter_mm, 1e-9)) or (p >= diameter_mm and (p - diameter_mm) <= 6.0) for p in perp_dims)
        is_cylinder = area_frac_ok or valid_perp

        if is_cylinder:
            # Snap diameter to common sizes if close
            try:
                common_diams = [10,12,15,16,18,20,22,25,28,30,32,35,40,45,50]
                closest_d = min(common_diams, key=lambda g: abs(g - diameter_mm))
                d_tol_abs = 1.5
                d_tol_rel = 0.04 * max(closest_d, 1.0)
                if abs(closest_d - diameter_mm) <= max(d_tol_abs, d_tol_rel):
                    diameter_mm = float(closest_d)
            except Exception:
                pass
            # Use selected diameter for stock volume instead of base radius
            volume_mm3 = math.pi * (0.25 * diameter_mm * diameter_mm) * length_mm
            dbg = {}
            try:
                dbg['axis_source'] = 'cylinder_faces'
                if profile_bins_debug is not None:
                    dbg['radial_profile_mm'] = profile_bins_debug
                if profile_meta is not None:
                    dbg['profile_meta'] = profile_meta
                dbg['circle_edge_max_r_mm'] = max_circle_r
                dbg['mesh_max_r_mm'] = mesh_max_r
            except Exception:
                pass
            result.update({
                'type': 'cylinder',
                'stock': {
                    'diameter_mm': _r2(diameter_mm),
                    'length_mm': _r2(length_mm),
                },
                'volume_mm3': volume_mm3,
                'debug': dbg,
            })
        else:
            # Treat as box if cylinder dims don't match bounding dims (e.g., holes in a block)
            if prefer_obb and obb_available:
                w, d, h = obb_dims
            else:
                w, d, h = (aabb_dx, aabb_dy, aabb_dz)
            # Prefer refined/measured thickness over extents when available
            try:
                if refined_thickness is not None and refined_thickness > 0:
                    h = refined_thickness
                # If histogram-based thickness is available, use it (more robust to small features)
                if 'hist_thickness' in locals() and hist_thickness is not None and hist_thickness > 0:
                    h = hist_thickness
                # Snap to common plate gauges if close; more permissive for plate-like parts, prefer snapping DOWN
                common_gauges = [5,6,8,10,12,15,18,20,22,25,28,30,32,35,40,45,50]
                closest = min(common_gauges, key=lambda g: abs(g - h))
                plate_like_pre = (h <= min(w, d) * 0.25) and (h <= 60.0)
                tol_abs = 2.0 if plate_like_pre else 0.8  # mm
                tol_rel = (0.05 if plate_like_pre else 0.03) * max(closest, 1.0)
                if plate_like_pre:
                    lower_pref_tol_abs = 2.6  # mm
                    g_lower_candidates = [g for g in common_gauges if g <= h]
                    if g_lower_candidates:
                        g_low = max(g_lower_candidates)
                        if (h - g_low) <= max(lower_pref_tol_abs, tol_rel):
                            h = float(g_low)
                        elif abs(closest - h) <= max(tol_abs, tol_rel):
                            h = float(closest)
                    else:
                        if abs(closest - h) <= max(tol_abs, tol_rel):
                            h = float(closest)
                else:
                    if abs(closest - h) <= max(tol_abs, tol_rel):
                        h = float(closest)
            except Exception:
                pass
            # Secondary fallback: infer cylinder from OBB/AABB if two dims are nearly equal
            try:
                dims = (w, d, h)
                # Identify potential cylinder: two smallest dims nearly equal; third is length
                srt = sorted(dims)
                dia_guess = (srt[0] + srt[1]) * 0.5
                len_guess = srt[2]
                eq_rel_tol = 0.03  # 3%
                eq_abs_tol = 1.5   # 1.5 mm
                if abs(srt[0] - srt[1]) <= max(eq_abs_tol, eq_rel_tol * max(dia_guess, 1e-9)):
                    # Reclassify as cylinder using inferred diameter/length
                    result.update({
                        'type': 'cylinder',
                        'stock': {
                            'diameter_mm': _r2(dia_guess),
                            'length_mm': _r2(len_guess),
                        },
                        'volume_mm3': math.pi * (0.25 * dia_guess * dia_guess) * len_guess,
                    })
                    # Skip box path by returning early
                    return result
            except Exception:
                pass
            # Normalize ordering so the smallest dimension is height/thickness
            dims_sorted = sorted([w, d, h])
            w, d, h = dims_sorted[1], dims_sorted[2], dims_sorted[0]
            # For plate-like parts, round width/depth to nearest mm for cleaner stock
            try:
                plate_like_post = (h <= min(w, d) * 0.25) and (h <= 60.0)
                if plate_like_post:
                    w = float(round(w))
                    d = float(round(d))
            except Exception:
                pass
            volume_mm3 = w * d * h
            result.update({
                'type': 'box',
                'stock': {
                    'width_mm': _r2(w),
                    'depth_mm': _r2(d),
                    'height_mm': _r2(h),
                },
                'volume_mm3': volume_mm3,
            })
    else:
        # Default to box using OBB when available, else AABB
        # Prefer face-aligned OBB extents (from planar axes), then OCCT OBB, then principal, else AABB
        if face_extents is not None:
            w, d, h = face_extents
            if face_dominant_thickness is not None:
                h = face_dominant_thickness
        elif prefer_obb and obb_available:
            w, d, h = obb_dims
        elif principal_extents is not None:
            w, d, h = principal_extents
        else:
            w, d, h = (aabb_dx, aabb_dy, aabb_dz)
        # Prefer refined/measured thickness over extents when available
        if refined_thickness is not None and refined_thickness > 0:
            h = refined_thickness
        # If histogram-based thickness is available, use it (more robust to small features)
        if 'hist_thickness' in locals() and hist_thickness is not None and hist_thickness > 0:
            h = hist_thickness
        # Snap to common plate gauges if close; more permissive for plate-like parts
        try:
            common_gauges = [5,6,8,10,12,15,18,20,22,25,28,30,32,35,40,45,50]
            closest = min(common_gauges, key=lambda g: abs(g - h))
            # Detect plate-like before sorting: thickness much smaller than other dims
            plate_like_pre = (h <= min(w, d) * 0.25) and (h <= 60.0)
            tol_abs = 2.0 if plate_like_pre else 0.8  # mm
            tol_rel = (0.05 if plate_like_pre else 0.03) * max(closest, 1.0)
            # Prefer snapping DOWN to the nearest lower gauge for plate-like parts
            if plate_like_pre:
                lower_pref_tol_abs = 2.6  # mm, more permissive downward preference
                g_lower_candidates = [g for g in common_gauges if g <= h]
                if g_lower_candidates:
                    g_low = max(g_lower_candidates)
                    if (h - g_low) <= max(lower_pref_tol_abs, tol_rel):
                        h = float(g_low)
                    elif abs(closest - h) <= max(tol_abs, tol_rel):
                        h = float(closest)
                else:
                    if abs(closest - h) <= max(tol_abs, tol_rel):
                        h = float(closest)
            else:
                if abs(closest - h) <= max(tol_abs, tol_rel):
                    h = float(closest)
        except Exception:
            pass
        # Normalize ordering so the smallest dimension is height/thickness
        dims_sorted = sorted([w, d, h])
        w, d, h = dims_sorted[1], dims_sorted[2], dims_sorted[0]
        # For plate-like parts, round width/depth to nearest mm for cleaner stock
        try:
            plate_like_post = (h <= min(w, d) * 0.25) and (h <= 60.0)
            if plate_like_post:
                w = float(round(w))
                d = float(round(d))
        except Exception:
            pass
        volume_mm3 = w * d * h
        result.update({
            'type': 'box',
            'stock': {
                'width_mm': _r2(w),
                'depth_mm': _r2(d),
                'height_mm': _r2(h),
            },
            'volume_mm3': volume_mm3,
        })

    return result


def compute_step_mass_properties(filepath):
    # Load and clean shape (includes solid filtering and unit heuristic)
    shape = _load_and_clean_step_shape(filepath)

    # Volume-based properties
    vprops = GProp_GProps()
    brepgprop_VolumeProperties(shape, vprops)
    volume = vprops.Mass()
    com = vprops.CentreOfMass()  # gp_Pnt
    inertia_mat = vprops.MatrixOfInertia()  # gp_Mat (relative to COM)
    pprops = vprops.PrincipalProperties()   # GProp_PrincipalProps

    # Surface area
    aprops = GProp_GProps()
    brepgprop_SurfaceProperties(shape, aprops)
    area = aprops.Mass()

    # Extract inertia tensor components (about COM)
    Ixx = inertia_mat.Value(1, 1)
    Iyy = inertia_mat.Value(2, 2)
    Izz = inertia_mat.Value(3, 3)
    Ixy = inertia_mat.Value(1, 2)
    Ixz = inertia_mat.Value(1, 3)
    Iyz = inertia_mat.Value(2, 3)

    # Principal moments and axes
    I1 = I2 = I3 = None
    v1 = v2 = v3 = None
    try:
        # Moments (principal)
        # PythonOCC exposes Moments() returning tuple (Ixx, Iyy, Izz) of principal moments
        I1, I2, I3 = pprops.Moments()
        # Axes are gp_Vec
        axis1 = pprops.FirstAxisOfInertia()
        axis2 = pprops.SecondAxisOfInertia()
        axis3 = pprops.ThirdAxisOfInertia()
        v1 = (axis1.X(), axis1.Y(), axis1.Z())
        v2 = (axis2.X(), axis2.Y(), axis2.Z())
        v3 = (axis3.X(), axis3.Y(), axis3.Z())
    except Exception:
        pass

    return {
        'volume_mm3': volume,
        'surface_area_mm2': area,
        'centre_of_mass_mm': {
            'x': com.X(), 'y': com.Y(), 'z': com.Z()
        },
        'inertia_tensor_mm5': {
            'Ixx': Ixx, 'Iyy': Iyy, 'Izz': Izz,
            'Ixy': Ixy, 'Ixz': Ixz, 'Iyz': Iyz
        },
        'principal_moments_mm5': {'I1': I1, 'I2': I2, 'I3': I3},
        'principal_axes': {'v1': v1, 'v2': v2, 'v3': v3}
    }


@app.route('/api/step/volume', methods=['POST'])
def api_step_volume():
    if not OCC_AVAILABLE:
        return jsonify({
            'error': 'pythonocc-core not installed',
            'hint': 'Install Python and run: python -m pip install pythonocc-core'
        }), 501
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    start = time.time()
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.step') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        volume = compute_step_volume(tmp_path)
        runtime_ms = int((time.time() - start) * 1000)
        os.unlink(tmp_path)
        return jsonify({
            'file': file.filename,
            'volume_mm3': volume,
            'runtime_ms': runtime_ms,
            'units': 'mm^3'
        })
    except Exception as e:
        app.logger.exception('api_step_volume failed')
        return jsonify({'error': str(e)}), 500


@app.route('/api/step/area', methods=['POST'])
def api_step_area():
    if not OCC_AVAILABLE:
        return jsonify({'error': 'pythonocc-core not installed'}), 501
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    start = time.time()
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.step') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        area = compute_step_area(tmp_path)
        runtime_ms = int((time.time() - start) * 1000)
        os.unlink(tmp_path)
        return jsonify({
            'file': file.filename,
            'surface_area_mm2': area,
            'runtime_ms': runtime_ms,
            'units': 'mm^2'
        })
    except Exception as e:
        app.logger.exception('api_step_area failed')
        return jsonify({'error': str(e)}), 500


@app.route('/api/step/mprops', methods=['POST'])
def api_step_mprops():
    if not OCC_AVAILABLE:
        return jsonify({'error': 'pythonocc-core not installed'}), 501
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    start = time.time()
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.step') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        props = compute_step_mass_properties(tmp_path)
        runtime_ms = int((time.time() - start) * 1000)
        os.unlink(tmp_path)
        props['file'] = file.filename
        props['runtime_ms'] = runtime_ms
        return jsonify(props)
    except Exception as e:
        app.logger.exception('api_step_mprops failed')
        return jsonify({'error': str(e)}), 500

@app.route('/api/step/stock', methods=['POST'])
def api_step_stock():
    if not OCC_AVAILABLE:
        return jsonify({'error': 'pythonocc-core not installed'}), 501
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    start = time.time()
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.step') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        stock = compute_step_stock_material(tmp_path)
        runtime_ms = int((time.time() - start) * 1000)
        os.unlink(tmp_path)
        stock['file'] = file.filename
        stock['runtime_ms'] = runtime_ms
        return jsonify(stock)
    except Exception as e:
        app.logger.exception('api_step_stock failed')
        return jsonify({'error': str(e)}), 500

@app.route('/api/save-excel', methods=['POST'])
def api_save_excel():
    """Browser fallback: save results to Excel 'Data' sheet (compatible with Electron behavior)."""
    try:
        payload = request.get_json(silent=True, force=True) or {}
        file_path = payload.get('filePath')
        headers = payload.get('headers')
        row = payload.get('row')
        if not file_path or not isinstance(headers, list) or not isinstance(row, list):
            return jsonify({'ok': False, 'error': 'Invalid payload'}), 400
        headers = [("Condition Note" if str(h) == "Sandblast Note" else h) for h in headers]

        import os
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)

        try:
            from openpyxl import Workbook, load_workbook
        except Exception as e:
            return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 501

        wb = None
        if os.path.exists(file_path):
            try:
                wb = load_workbook(file_path)
            except Exception as e:
                # Common case: file locked by Excel
                msg = str(e)
                if 'in use' in msg or 'Permission denied' in msg or 'Access is denied' in msg:
                    return jsonify({'ok': False, 'error': 'EBUSY: file is open/locked'}), 423
                return jsonify({'ok': False, 'error': msg}), 500
        if wb is None:
            wb = Workbook()
        sheet = wb['Data'] if 'Data' in wb.sheetnames else wb.create_sheet('Data', 0)

        # sanitize existing header row
        if sheet.max_row >= 1:
            r1 = sheet[1]
            for c in r1:
                if c.value == "Sandblast Note":
                    c.value = "Condition Note"
        # Add headers if first row empty
        if sheet.max_row == 1 and sheet.max_column == 1 and (sheet.cell(row=1, column=1).value is None):
            sheet.append(headers)
        # Always append data row
        sheet.append(row)

        try:
            wb.save(file_path)
        except Exception as e:
            msg = str(e)
            if 'in use' in msg or 'Permission denied' in msg or 'Access is denied' in msg:
                return jsonify({'ok': False, 'error': 'EBUSY: file is open/locked'}), 423
            return jsonify({'ok': False, 'error': msg}), 500
        return jsonify({'ok': True})
    except Exception as e:
        app.logger.exception('api_save_excel failed')
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/open-file', methods=['POST', 'OPTIONS'])
def api_open_file():
    """Open the specified file if it exists; otherwise open its directory.
    Matches browser fallback behavior used when Electron APIs are unavailable.
    """
    try:
        payload = request.get_json(silent=True, force=True) or {}
        file_path = payload.get('filePath')
        if not file_path or not isinstance(file_path, str):
            return jsonify({'ok': False, 'error': 'Invalid filePath'}), 400

        exists = os.path.exists(file_path)
        if exists:
            try:
                if os.name == 'nt':
                    os.startfile(file_path)  # type: ignore[attr-defined]
                else:
                    subprocess.Popen(['xdg-open', file_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return jsonify({'ok': True})
            except Exception as e:
                app.logger.exception('open-file failed')
                return jsonify({'ok': False, 'error': str(e)}), 500
        else:
            dir_path = os.path.dirname(file_path) or '.'
            try:
                if os.name == 'nt':
                    subprocess.Popen(['explorer.exe', dir_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                else:
                    subprocess.Popen(['xdg-open', dir_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return jsonify({'ok': False, 'error': 'File not found. Opened directory instead.'})
            except Exception as e:
                app.logger.exception('open-file directory fallback failed')
                return jsonify({'ok': False, 'error': str(e)}), 500
    except Exception as e:
        app.logger.exception('api_open_file failed')
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/calibration/reference', methods=['GET'])
def api_calibration_reference():
    """Return the saved calibration reference marker with method and result."""
    try:
        if not os.path.exists(CALIBRATION_PATH):
            return jsonify({'ok': False, 'error': 'calibration_reference.json not found'}), 404
        with open(CALIBRATION_PATH, 'r', encoding='utf-8') as f:
            import json as _json
            data = _json.load(f)
        return jsonify({'ok': True, 'data': data})
    except Exception as e:
        app.logger.exception('api_calibration_reference failed')
        return jsonify({'ok': False, 'error': str(e)}), 500

# Fallback for GET under /api/* to guarantee /api/health responds 200
@app.route('/api/<path:subpath>', methods=['GET', 'HEAD'])
def api_get_fallback(subpath):
    try:
        sp = str(subpath).rstrip('/')
    except Exception:
        sp = subpath
    if sp == 'health':
        try:
            app.logger.info("FALLBACK /api/<subpath> matched 'health' -> 200")
        except Exception:
            pass
        return jsonify({
            'ok': True,
            'status': 'running',
            'occ_available': bool(OCC_AVAILABLE),
            'time': time.time()
        })
    return jsonify({'ok': False, 'error': 'Unknown endpoint'}), 404

# Static catch-all must be declared AFTER API routes to avoid shadowing them
@app.route('/<path:path>')
def serve_file(path):
    try:
        app.logger.info(f"serve_file route called path='{path}'")
    except Exception:
        pass
    # Special-case: ensure /api/health always returns 200 even if catch-all matched
    try:
        p = str(path).rstrip('/')
    except Exception:
        p = path
    if p == 'api/health':
        return jsonify({
            'ok': True,
            'status': 'running',
            'occ_available': bool(OCC_AVAILABLE),
            'time': time.time()
        })
    return send_from_directory(BASE_DIR, path)

if __name__ == '__main__':
    # Determine port:
    # - Packaged (frozen) EXE uses 5001 for production
    # - Dev uses FLASK_PORT (default 5002) to avoid collisions with packaged EXE
    try:
        _is_frozen = getattr(sys, 'frozen', False)
    except Exception:
        _is_frozen = False
    if _is_frozen:
        _port = 5001
    else:
        try:
            _port = int(os.environ.get('FLASK_PORT', '5002'))
        except Exception:
            _port = 5002
    app.logger.info(f"Starting CNC Costify AI Flask API on port {_port} | log_dir={LOG_DIR}")
    try:
        app.logger.info(f"URL Map: {app.url_map}")
    except Exception:
        pass
    try:
        funcs = app.before_request_funcs.get(None, [])
        app.logger.info(f"Before request funcs: {len(funcs)} -> {[getattr(f, '__name__', str(f)) for f in funcs]}")
        app.logger.info(f"View functions: {list(app.view_functions.keys())}")
    except Exception:
        pass
    # Explicitly prevent opening external browser on start in all modes
    # The Electron app is responsible for presenting the UI.
    app.run(debug=False, port=_port)
